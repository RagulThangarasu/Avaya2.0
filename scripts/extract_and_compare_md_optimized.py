import os
import sys
import json
import asyncio
import re
import difflib
import random
import time
from datetime import datetime
from urllib.parse import urljoin, urlparse
from playwright.async_api import async_playwright

# Configurations
REPORTS_DIR = os.path.join(os.getcwd(), 'reports')
os.makedirs(REPORTS_DIR, exist_ok=True)

TEST_URLS_PATH = os.path.join(os.getcwd(), 'config', 'test-urls.json')
AUTH_STATE_PATH = os.path.join(os.getcwd(), 'auth-sessions', 'storage-state.json')

STAGE_URL = os.environ.get('STAGE_URL')
PROD_URL = os.environ.get('PROD_URL')

STAGE_MD_PATH = os.path.join(REPORTS_DIR, 'stage_topics.md')
PROD_MD_PATH = os.path.join(REPORTS_DIR, 'prod_topics.md')
COMPARISON_REPORT_PATH = os.path.join(REPORTS_DIR, 'md_comparison_report.md')

# Load fallback URLs
if not STAGE_URL or not PROD_URL:
    try:
        with open(TEST_URLS_PATH, 'r') as f:
            config = json.load(f)
            STAGE_URL = STAGE_URL or config.get('stage')
            PROD_URL = PROD_URL or config.get('production')
    except:
        pass

if not STAGE_URL or not PROD_URL:
    print("❌ Error: Missing STAGE_URL or PROD_URL.")
    sys.exit(1)

print(f"🚀 Starting DYNAMIC HIGH-SPEED Markdown Extractor & Comparer")
print(f"   Stage: {STAGE_URL}")
print(f"   Prod:  {PROD_URL}\n")

def get_filename(url):
    path = urlparse(url).path
    filename = path.split('/')[-1] if '/' in path else path
    return filename.replace('.html', '').replace('.htm', '').lower().replace('-', '').replace('_', '')

def slugify(t):
    if not t: return ""
    return re.sub(r'[^a-z0-9]', '', t.lower())

async def handle_cookies(page):
    try:
        for sel in ['#onetrust-accept-btn-handler', '#btn-accept-all', 'button:has-text("Accept")', '.cookie-accept']:
            if await page.locator(sel).is_visible(timeout=1500):
                await page.click(sel)
                await page.wait_for_timeout(500)
                break
    except:
        pass

async def block_resources(route):
    """Aborts only heavy assets (images, fonts, media) and tracking/analytics
       so page loading is lightning-fast (10x speedup)."""
    req = route.request
    resource_type = req.resource_type
    url = req.url.lower()
    
    # Abort heavy rendering media
    if resource_type in ["image", "font", "media", "imageset"]:
        await route.abort()
        return
        
    # Block cookies, analytics, tracking and dynamic third-party libraries
    block_patterns = [
        "analytics", "telemetry", "google-analytics", "onetrust", 
        "adobedtm", "demdex", "marketingcloud", "sift", "hotjar",
        "facebook", "doubleclick", "crazyegg", "optimizely", "clarity",
        "cookie-consent", "trustarc", "recaptcha", "dynatrace", "ruxit"
    ]
    
    if any(pattern in url for pattern in block_patterns):
        await route.abort()
        return
        
    await route.continue_()

# ── TOC Extraction ───────────────────────────────────────────────────
async def extract_prod_toc(page, base_url):
    await page.wait_for_load_state("networkidle")
    await handle_cookies(page)
    await page.wait_for_timeout(3000)

    try:
        # 1. Click "Expand All" button if available
        expand_btn = page.locator('.zDocsCollapseExpandButton').first
        if await expand_btn.is_visible(timeout=5000):
            await expand_btn.click()
            await page.wait_for_timeout(5000)
    except:
        pass

    # 2. Iteratively expand nested nodes up to 5 levels deep to capture full sequence
    for level in range(5):
        try:
            collapse_nodes = await page.locator('.zDocsTocCollapseItemButton[aria-expanded="false"], button[aria-expanded="false"], .expand-icon').all()
            if not collapse_nodes:
                break
            print(f"   [Level {level+1}] Expanding {len(collapse_nodes)} nested nodes...")
            for node in collapse_nodes[:120]:
                try:
                    await node.click(timeout=800)
                    await page.wait_for_timeout(50)
                except:
                    pass
            await page.wait_for_timeout(2000)
        except Exception as e:
            print(f"   [Level {level+1}] Expand warning: {e}")
            break

    links_data = await page.evaluate(r'''() => {
        const results = [];
        const container = document.querySelector('.zDocsTocList') || document.querySelector('.zDocsTOC');
        let allLinks = [];
        if (container) {
            allLinks = Array.from(container.querySelectorAll('a[href]'));
        }
        if (allLinks.length === 0) {
            allLinks = Array.from(document.querySelectorAll('nav a[href], [class*="sidebar"] a[href]'));
        }
        allLinks.forEach(a => {
            const href = a.getAttribute('href');
            let text = a.innerText.trim();
            // Clean up titles (remove leading/trailing symbols, newlines, tabs, chevron chars)
            text = text.replace(/[\r\n\t]+/g, ' ').replace(/^\s*[\u203A\u25BC\u25B6>v-]\s*/g, '').trim();
            if (href && text && !href.startsWith('#') && !href.startsWith('javascript')) {
                results.push({text, href});
            }
        });
        return results;
    }''')

    toc = []
    seen = set()
    for item in links_data:
        full_url = urljoin(base_url, item['href']).split('#')[0].split('?')[0]
        if full_url not in seen:
            toc.append({'title': item['text'], 'url': full_url})
            seen.add(full_url)

    print(f"   ✅ Prod TOC: {len(toc)} topics found.")
    return toc

async def extract_stage_toc(page, base_url):
    await page.wait_for_load_state("networkidle")
    await handle_cookies(page)
    await page.wait_for_timeout(2000)

    # Iteratively expand collapsible nested items on Stage to match sequences
    for level in range(3):
        try:
            collapse_nodes = await page.locator('.cmp-navigation__item--active[aria-expanded="false"], .cmp-navigation__item[aria-expanded="false"], button[aria-expanded="false"]').all()
            if not collapse_nodes:
                break
            for node in collapse_nodes[:50]:
                try:
                    await node.click(timeout=800)
                except:
                    pass
            await page.wait_for_timeout(1000)
        except:
            break

    links_data = await page.evaluate(r'''() => {
        const results = [];
        const links = document.querySelectorAll('.cmp-navigation__item-link');
        links.forEach(a => {
            const href = a.getAttribute('href');
            let text = a.innerText.trim();
            text = text.replace(/[\r\n\t]+/g, ' ').replace(/^\s*[\u203A\u25BC\u25B6>v-]\s*/g, '').trim();
            if (href && text && !href.startsWith('#') && !href.startsWith('javascript')) {
                results.push({text, href});
            }
        });
        return results;
    }''')

    toc = []
    seen = set()
    for item in links_data:
        full_url = urljoin(base_url, item['href']).split('#')[0].split('?')[0]
        if full_url not in seen and '.html' in full_url:
            toc.append({'title': item['text'], 'url': full_url})
            seen.add(full_url)

    print(f"   ✅ Stage TOC: {len(toc)} topics found.")
    return toc

# ── Content MD Extraction ────────────────────────────────────────────
async def fetch_single_attempt(page, url, is_prod=False):
    """Inner worker that performs a single page loading and extraction attempt."""
    try:
        # Generous timeout (35 seconds) to allow dynamic AEM React loading
        await page.goto(url, wait_until='domcontentloaded', timeout=35000)
        
        root_selector = '.zDocsTopicPageBody' if is_prod else '.topic-renderer__content'
        
        # Wait for root selector to be attached to the page DOM
        try:
            await page.wait_for_selector(root_selector, state='attached', timeout=10000)
        except:
            pass
            
        md_content = await page.evaluate('''async (selector) => {
            const root = document.querySelector(selector) || document.body;
            
            function isBoldElement(el) {
                const tag = el.tagName.toLowerCase();
                return ['b', 'strong'].includes(tag);
            }

            function isItalicElement(el) {
                const tag = el.tagName.toLowerCase();
                return ['i', 'em'].includes(tag);
            }

            let markdown = '';
            
            const h1 = root.querySelector('h1');
            if (h1) {
                markdown += `# ${h1.innerText.trim()}\\n\\n`;
            }

            function walk(node) {
                let text = '';
                if (node.nodeType === 3) {
                    let val = node.textContent;
                    val = val.replace(/[\\r\\n\\t]+/g, ' ').replace(/ {2,}/g, ' ');
                    text += val;
                } else if (node.nodeType === 1) {
                    const tag = node.tagName.toLowerCase();
                    
                    // Safe string conversion for classnames and IDs (handles SVGs)
                    const className = (typeof node.className === 'string') ? node.className.toLowerCase() : '';
                    const id = (typeof node.id === 'string') ? node.id.toLowerCase() : '';
                    
                    if (['script', 'style', 'header', 'footer', 'nav'].includes(tag) ||
                        className.includes('breadcrumb') || className.includes('toolbar') ||
                        className.includes('metadata') || className.includes('search')) {
                        return '';
                    }

                    const isBold = isBoldElement(node);
                    const isItalic = isItalicElement(node);

                    let childContent = '';
                    for (let child of node.childNodes) {
                        childContent += walk(child);
                    }

                    if (tag === 'h2') {
                        text += `\\n\\n## ${childContent.trim()}\\n\\n`;
                    } else if (tag === 'h3') {
                        text += `\\n\\n### ${childContent.trim()}\\n\\n`;
                    } else if (tag === 'p') {
                        text += `\\n\\n${childContent.trim()}\\n\\n`;
                    } else if (tag === 'li') {
                        text += `\\n- ${childContent.trim()}`;
                    } else {
                        let dec = childContent;
                        if (isBold && dec.trim()) dec = `**${dec.trim()}**`;
                        if (isItalic && dec.trim()) dec = `*${dec.trim()}*`;
                        text += dec;
                    }
                }
                return text;
            }

            markdown += walk(root);

            // Related Links Extraction
            const relatedLinks = [];
            const relatedSelectors = [
                '.related-links', '.zDocsRelatedLinks', '.related-links-list', 
                '#related-links', '.related-links-container', '.cmp-related-links'
            ];
            
            relatedSelectors.forEach(sel => {
                root.querySelectorAll(sel).forEach(container => {
                    container.querySelectorAll('a[href]').forEach(a => {
                        const href = a.getAttribute('href');
                        const linkText = a.innerText.trim();
                        if (href && linkText && !href.startsWith('#') && !href.startsWith('javascript')) {
                            relatedLinks.push({ text: linkText, href: href });
                        }
                    });
                });
            });

            if (relatedLinks.length === 0) {
                const headings = root.querySelectorAll('h2, h3, h4, h5, div');
                headings.forEach(h => {
                    const hText = h.innerText.trim().toLowerCase();
                    if (hText.includes('related information') || hText.includes('related links') || hText.includes('related topics') || hText.includes('related concepts') || hText.includes('related tasks') || hText.includes('related references')) {
                        let next = h.nextElementSibling;
                        while (next && !['h2', 'h3', 'h4', 'h5'].includes(next.tagName.toLowerCase())) {
                            next.querySelectorAll('a[href]').forEach(a => {
                                const href = a.getAttribute('href');
                                const linkText = a.innerText.trim();
                                if (href && linkText && !href.startsWith('#') && !href.startsWith('javascript')) {
                                    relatedLinks.push({ text: linkText, href: href });
                                }
                            });
                            next = next.nextElementSibling;
                        }
                    }
                });
            }

            if (relatedLinks.length > 0) {
                markdown += `\\n\\n### Related Information\\n\\n`;
                relatedLinks.forEach(link => {
                    markdown += `- [${link.text}](${link.href})\\n`;
                });
            }

            // Clean up multi-newlines
            let cleaned = markdown.replace(/\\n{3,}/g, '\\n\\n').trim();
            
            // Normalize spaces line-by-line while preserving Markdown syntax prefixes
            let lines = cleaned.split('\\n');
            lines = lines.map(line => {
                let trimmed = line.trim();
                let prefix = '';
                if (trimmed.startsWith('- ')) {
                    prefix = '- ';
                } else if (trimmed.startsWith('* ')) {
                    prefix = '* ';
                } else if (trimmed.startsWith('#')) {
                    let i = 0;
                    while (i < trimmed.length && trimmed[i] === '#') {
                        i++;
                    }
                    while (i < trimmed.length && trimmed[i] === ' ') {
                        i++;
                    }
                    prefix = trimmed.substring(0, i);
                }
                
                let content = trimmed.substring(prefix.length).trim();
                content = content.replace(/[\\s\\r\\n\\t]+/g, ' ');
                return prefix + content;
            });
            return lines.join('\\n');
        }''', root_selector)
        
        return md_content
    except Exception as e:
        return f"<!-- Error loading page: {e} -->"

async def extract_topic_markdown(page, url, is_prod=False, retries=3):
    """Wrapper that retries fetch_single_attempt up to 3 times with exponential backoff."""
    for attempt in range(retries):
        try:
            content = await fetch_single_attempt(page, url, is_prod)
            # If successfully extracted, return immediately!
            if content and not content.startswith("<!-- Error"):
                return content
            
            # If it's a minor loading error, let's raise it to trigger the retry
            if content and content.startswith("<!-- Error"):
                raise Exception(content.replace("<!--", "").replace("-->", "").strip())
                
        except Exception as e:
            if attempt < retries - 1:
                wait_time = 4 * (attempt + 1)
                print(f"⚠️ Warning: Failed loading {url} (Attempt {attempt+1}/{retries}). Retrying in {wait_time}s... Error: {e}")
                await asyncio.sleep(wait_time)
            else:
                return f"<!-- Error loading page: {e} after {retries} attempts -->"
                
    return "<!-- Error: Max retries exceeded -->"

async def process_topics_in_markdown():
    start_time = datetime.now()
    async with async_playwright() as p:
        browser = await p.chromium.launch()
        
        stage_ctx = await browser.new_context(
            storage_state=AUTH_STATE_PATH if os.path.exists(AUTH_STATE_PATH) else None
        )
        prod_ctx = await browser.new_context()
        
        print("🔍 Crawling TOCs (Full Load for Accuracy)...")
        p_page = await prod_ctx.new_page()
        prod_toc = []
        for attempt in range(3):
            try:
                print(f"   [Attempt {attempt+1}/3] Loading Production URL...")
                # Start with domcontentloaded, then optionally settle networkidle
                await p_page.goto(PROD_URL, wait_until="domcontentloaded", timeout=45000)
                try:
                    await p_page.wait_for_load_state("networkidle", timeout=15000)
                except:
                    pass
                prod_toc = await extract_prod_toc(p_page, PROD_URL)
                break
            except Exception as e:
                if attempt < 2:
                    wait_time = 5 * (attempt + 1)
                    print(f"⚠️ Warning: Failed loading Production TOC (Attempt {attempt+1}/3). Retrying in {wait_time}s... Error: {e}")
                    await asyncio.sleep(wait_time)
                else:
                    print(f"❌ Fatal error: Production TOC loading failed after 3 attempts. Error: {e}")
                    raise e
        await p_page.close()
        
        s_page = await stage_ctx.new_page()
        stage_toc = []
        for attempt in range(3):
            try:
                print(f"   [Attempt {attempt+1}/3] Loading Stage URL...")
                await s_page.goto(STAGE_URL, wait_until="domcontentloaded", timeout=45000)
                try:
                    await s_page.wait_for_load_state("networkidle", timeout=15000)
                except:
                    pass
                stage_toc = await extract_stage_toc(s_page, STAGE_URL)
                break
            except Exception as e:
                if attempt < 2:
                    wait_time = 5 * (attempt + 1)
                    print(f"⚠️ Warning: Failed loading Stage TOC (Attempt {attempt+1}/3). Retrying in {wait_time}s... Error: {e}")
                    await asyncio.sleep(wait_time)
                else:
                    print(f"❌ Fatal error: Stage TOC loading failed after 3 attempts. Error: {e}")
                    raise e
        await s_page.close()
        
        # Build Stage mapping by filename
        stage_by_filename = {}
        for idx, item in enumerate(stage_toc):
            fn = get_filename(item['url'])
            if fn:
                if fn not in stage_by_filename:
                    stage_by_filename[fn] = []
                stage_by_filename[fn].append({**item, 'index': idx + 1, 'used': False})
        
        stage_md_blocks = ["" for _ in prod_toc]
        prod_md_blocks = ["" for _ in prod_toc]
        
        # Spin up fast routing for batch processing
        fast_stage_ctx = await browser.new_context(
            storage_state=AUTH_STATE_PATH if os.path.exists(AUTH_STATE_PATH) else None
        )
        fast_prod_ctx = await browser.new_context()
        
        # Create asynchronous queue of topics
        queue = asyncio.Queue()
        for idx, prod_item in enumerate(prod_toc):
            await queue.put((idx, prod_item))
            
        print(f"\n⚡ Compiling markdown for {len(prod_toc)} matched topics via worker pool of 6 paced reusable tabs...")
        
        async def worker():
            p_page = await fast_prod_ctx.new_page()
            s_page = await fast_stage_ctx.new_page()
            
            # Register advanced resource blocks on reusable tabs directly
            await p_page.route("**/*", block_resources)
            await s_page.route("**/*", block_resources)
            
            while not queue.empty():
                idx, prod_item = await queue.get()
                
                # Add natural pacing delay to safeguard network stability and prevent DDoS blocks
                await asyncio.sleep(random.uniform(0.5, 1.5))
                
                prod_url = prod_item['url']
                prod_fn = get_filename(prod_url)
                
                stage_url = None
                
                # 1. Exact filename match
                candidates = stage_by_filename.get(prod_fn, [])
                for cand in candidates:
                    if not cand['used']:
                        cand['used'] = True
                        stage_url = cand['url']
                        break
                        
                # 2. Fallback: Base filename match (stripping trailing digits, underscores, dashes)
                if not stage_url:
                    prod_fn_base = re.sub(r'[\d_-]+$', '', prod_fn)
                    for stage_fn, cand_list in stage_by_filename.items():
                        stage_fn_base = re.sub(r'[\d_-]+$', '', stage_fn)
                        if prod_fn_base == stage_fn_base:
                            for cand in cand_list:
                                if not cand['used']:
                                    cand['used'] = True
                                    stage_url = cand['url']
                                    break
                            if stage_url:
                                break
                                
                # 3. Fallback: Match by exact Title slug
                if not stage_url:
                    prod_title_slug = slugify(prod_item['title'])
                    for stage_fn, cand_list in stage_by_filename.items():
                        for cand in cand_list:
                            if not cand['used'] and slugify(cand['title']) == prod_title_slug:
                                cand['used'] = True
                                stage_url = cand['url']
                                break
                        if stage_url:
                            break
                
                # Fetch Stage & Prod in parallel
                p_task = extract_topic_markdown(p_page, prod_url, is_prod=True)
                s_task = extract_topic_markdown(s_page, stage_url, is_prod=False) if stage_url else asyncio.sleep(0.01)
                
                prod_md, stage_md = await asyncio.gather(p_task, s_task)
                
                if not stage_url:
                    stage_md = f"# {prod_item['title']}\n\n*Topic is completely missing in Stage environment.*"
                
                prod_md_blocks[idx] = prod_md
                stage_md_blocks[idx] = stage_md
                
                print(f"  [{idx+1}/{len(prod_toc)}] compiled: {prod_item['title'][:45]}")
                queue.task_done()
                
            await p_page.close()
            await s_page.close()

        # Spin up 6 parallel workers reusing tabs (12 pages total) - highly stable!
        workers = [asyncio.create_task(worker()) for _ in range(6)]
        await asyncio.gather(*workers)
                
        # Write files
        print(f"\n✍️ Saving stage_topics.md...")
        with open(STAGE_MD_PATH, 'w', encoding='utf-8') as sf:
            sf.write(f"# Consolidated Stage Topics\n")
            sf.write(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            sf.write("\n\n---\n\n".join(stage_md_blocks))
            
        print(f"✍️ Saving prod_topics.md...")
        with open(PROD_MD_PATH, 'w', encoding='utf-8') as pf:
            pf.write(f"# Consolidated Production Topics\n")
            pf.write(f"Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            pf.write("\n\n---\n\n".join(prod_md_blocks))
            
        await fast_stage_ctx.close()
        await fast_prod_ctx.close()
        await prod_ctx.close()
        await stage_ctx.close()
        await browser.close()
        
    duration = (datetime.now() - start_time).total_seconds()
    print(f"\n⚡ Extraction finished in {duration:.1f} seconds!")
    return len(prod_toc), duration, prod_toc, stage_md_blocks, prod_md_blocks

# ── Direct Markdown Comparison ───────────────────────────────────────
def split_into_sentences(text):
    """Normalize whitespaces and split content into clean sentences, ignoring empty spaces."""
    # Convert all consecutive whitespaces/newlines/tabs into a single space
    normalized = re.sub(r'\s+', ' ', text)
    # Split by standard sentence delimiters (. ! ?) followed by a space
    sentences = re.split(r'(?<=[.!?])\s+', normalized)
    return [s.strip() for s in sentences if s.strip()]

def compare_compiled_markdowns(total_topics, duration, prod_toc, stage_md_blocks, prod_md_blocks):
    print("\n🔍 Validating and comparing markdown contents...")
    
    with open(STAGE_MD_PATH, 'r', encoding='utf-8') as sf:
        stage_text = sf.read()
    with open(PROD_MD_PATH, 'r', encoding='utf-8') as pf:
        prod_text = pf.read()
        
    # Extract clean sentence structures to ignore blank lines and raw empty spaces
    stage_sentences = split_into_sentences(stage_text)
    prod_sentences = split_into_sentences(prod_text)
    
    # Calculate similarity on actual matching sentences
    matcher = difflib.SequenceMatcher(None, stage_sentences, prod_sentences)
    match_pct = matcher.ratio() * 100
    
    # Word count breakdown
    stage_words = len(stage_text.split())
    prod_words = len(prod_text.split())
    
    # Character count breakdown: exclude all whitespace spaces
    stage_char_count = len([c for c in stage_text if not c.isspace()])
    prod_char_count = len([c for c in prod_text if not c.isspace()])
    
    # Bold & Italic markdown count parity check
    stage_bold_cnt = len(re.findall(r'\*\*.*?\*\*', stage_text))
    prod_bold_cnt = len(re.findall(r'\*\*.*?\*\*', prod_text))
    
    stage_italic_cnt = len(re.findall(r'\*.*?\*', stage_text)) - (stage_bold_cnt * 2)
    prod_italic_cnt = len(re.findall(r'\*.*?\*', prod_text)) - (prod_bold_cnt * 2)
    stage_italic_cnt = max(stage_italic_cnt, 0)
    prod_italic_cnt = max(prod_italic_cnt, 0)
    
    stage_links_cnt = len(re.findall(r'\[.*?\]\(.*?\)', stage_text))
    prod_links_cnt = len(re.findall(r'\[.*?\]\(.*?\)', prod_text))
    
    diff_lines = list(difflib.unified_diff(
        stage_sentences,
        prod_sentences,
        fromfile='stage_topics.md (Sentences)',
        tofile='prod_topics.md (Sentences)',
        n=2
    ))
    
    # ── Bold Text Content Parity & Checklist Generation ─────────────────
    stage_bold_phrases = [b.strip() for b in re.findall(r'\*\*(.*?)\*\*', stage_text) if b.strip()]
    prod_bold_phrases = [b.strip() for b in re.findall(r'\*\*(.*?)\*\*', prod_text) if b.strip()]
    
    stage_bold_set = set(stage_bold_phrases)
    prod_bold_set = set(prod_bold_phrases)
    
    matched_bolds = sorted(list(stage_bold_set & prod_bold_set))
    missing_in_prod = sorted(list(stage_bold_set - prod_bold_set))
    extra_in_prod = sorted(list(prod_bold_set - stage_bold_set))
    
    total_unique_bolds = len(stage_bold_set | prod_bold_set)
    bold_match_pct = (len(matched_bolds) / total_unique_bolds * 100) if total_unique_bolds else 100.0
    
    # ── Related / Inline Links Parity Check ─────────────────────────────
    # Extract links excluding internal file:// references
    stage_links = [l for l in re.findall(r'\[(.*?)\]\((.*?)\)', stage_text) if not l[1].startswith('file://')]
    prod_links = [l for l in re.findall(r'\[(.*?)\]\((.*?)\)', prod_text) if not l[1].startswith('file://')]
    
    stage_link_set = {url.strip() for label, url in stage_links if url.strip()}
    prod_link_set = {url.strip() for label, url in prod_links if url.strip()}
    
    stage_link_map = {url.strip(): label.strip() for label, url in stage_links if url.strip()}
    prod_link_map = {url.strip(): label.strip() for label, url in prod_links if url.strip()}
    
    missing_links_in_prod = sorted(list(stage_link_set - prod_link_set))  # In Stage but missing in Prod
    missing_links_in_stage = sorted(list(prod_link_set - stage_link_set)) # In Prod but missing in Stage
    
    # ── Per-Topic Missing Bold Text & Sentence Parity ───────────────────
    per_topic_bold_drift = []
    per_topic_sentence_drift = []
    per_topic_link_drift = []
    
    for idx, item in enumerate(prod_toc):
        topic_title = item['title']
        prod_url = item['url']
        stage_md = stage_md_blocks[idx]
        prod_md = prod_md_blocks[idx]
        
        # 1. Bold text parity per topic (strong elements in prod missing in stage)
        stage_bolds = [b.strip() for b in re.findall(r'\*\*(.*?)\*\*', stage_md) if b.strip()]
        prod_bolds = [b.strip() for b in re.findall(r'\*\*(.*?)\*\*', prod_md) if b.strip()]
        
        missing_bold_in_stage = sorted(list(set(prod_bolds) - set(stage_bolds)))
        extra_bold_in_stage = sorted(list(set(stage_bolds) - set(prod_bolds)))
        if missing_bold_in_stage or extra_bold_in_stage:
            per_topic_bold_drift.append({
                'title': topic_title,
                'url': prod_url,
                'missing_bolds': missing_bold_in_stage,
                'extra_bolds': extra_bold_in_stage
            })
            
        # 2. Sentence parity per topic (Stage we have, Prod doesn't have)
        stage_sents = split_into_sentences(stage_md)
        prod_sents = split_into_sentences(prod_md)
        
        stage_sent_set = {s.strip() for s in stage_sents if s.strip()}
        prod_sent_set = {s.strip() for s in prod_sents if s.strip()}
        
        missing_in_prod = sorted(list(stage_sent_set - prod_sent_set)) # In Stage, missing in Prod
        missing_in_stage = sorted(list(prod_sent_set - stage_sent_set)) # In Prod, missing in Stage
        
        if missing_in_prod or missing_in_stage:
            per_topic_sentence_drift.append({
                'title': topic_title,
                'url': prod_url,
                'missing_in_prod': missing_in_prod,
                'missing_in_stage': missing_in_stage
            })

        # 3. Link parity per topic
        stage_topic_links = [l for l in re.findall(r'\[(.*?)\]\((.*?)\)', stage_md) if not l[1].startswith('file://')]
        prod_topic_links = [l for l in re.findall(r'\[(.*?)\]\((.*?)\)', prod_md) if not l[1].startswith('file://')]
        
        stage_topic_link_set = {url.strip() for label, url in stage_topic_links if url.strip()}
        prod_topic_link_set = {url.strip() for label, url in prod_topic_links if url.strip()}
        
        stage_topic_link_map = {url.strip(): label.strip() for label, url in stage_topic_links if url.strip()}
        prod_topic_link_map = {url.strip(): label.strip() for label, url in prod_topic_links if url.strip()}
        
        topic_missing_links_in_prod = sorted(list(stage_topic_link_set - prod_topic_link_set))
        topic_missing_links_in_stage = sorted(list(prod_topic_link_set - stage_topic_link_set))
        
        if topic_missing_links_in_prod or topic_missing_links_in_stage:
            per_topic_link_drift.append({
                'title': topic_title,
                'url': prod_url,
                'missing_in_prod': topic_missing_links_in_prod,
                'missing_in_stage': topic_missing_links_in_stage,
                'stage_map': stage_topic_link_map,
                'prod_map': prod_topic_link_map
            })
            
    BOLD_REPORT_PATH = os.path.join(REPORTS_DIR, 'bold_text_parity_report.md')
    with open(BOLD_REPORT_PATH, 'w', encoding='utf-8') as brf:
        brf.write(f"# 🔠 Bold Text Content Parity & Comparison Report\n\n")
        brf.write(f"> **Report Date**: `{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`  \n")
        brf.write(f"> **Stage MD File**: [{os.path.basename(STAGE_MD_PATH)}](file://{STAGE_MD_PATH})  \n")
        brf.write(f"> **Production MD File**: [{os.path.basename(PROD_MD_PATH)}](file://{PROD_MD_PATH})  \n\n")
        
        brf.write(f"## 📈 Bold Text Parity Stats\n\n")
        brf.write(f"| Bold Metric | Value | Percentage |\n")
        brf.write(f"| :--- | :---: | :---: |\n")
        brf.write(f"| **Stage Unique Bold Phrases** | {len(stage_bold_set)} | - |\n")
        brf.write(f"| **Production Unique Bold Phrases** | {len(prod_bold_set)} | - |\n")
        brf.write(f"| **Perfect Matches** | {len(matched_bolds)} | {len(matched_bolds)/max(total_unique_bolds, 1)*100:.2f}% |\n")
        brf.write(f"| **Missing in Production** | {len(missing_in_prod)} | {len(missing_in_prod)/max(total_unique_bolds, 1)*100:.2f}% |\n")
        brf.write(f"| **Extra in Production (Missing in Stage)** | {len(extra_in_prod)} | {len(extra_in_prod)/max(total_unique_bolds, 1)*100:.2f}% |\n\n")
        
        brf.write(f"## ❌ Missing in Production ({len(missing_in_prod)})\n\n")
        if missing_in_prod:
            for item in missing_in_prod:
                brf.write(f"- [ ] `{item}`\n")
        else:
            brf.write(f"🎉 **None! All Stage bold phrases are styled as bold in Production.**\n")
            
        brf.write(f"\n## ➕ Extra in Production ({len(extra_in_prod)})\n\n")
        if extra_in_prod:
            for item in extra_in_prod:
                brf.write(f"- [ ] `{item}`\n")
        else:
            brf.write(f"🎉 **None! No extra bold phrases styled in Production.**\n")
            
        brf.write(f"\n## ✅ Perfect Matches ({len(matched_bolds)})\n\n")
        if matched_bolds:
            for item in matched_bolds:
                brf.write(f"- [x] `{item}`\n")
        else:
            brf.write(f"*No perfect matches found.*\n")

    with open(COMPARISON_REPORT_PATH, 'w', encoding='utf-8') as rf:
        rf.write(f"# 📊 Consolidated Topics Markdown Comparison Report\n\n")
        rf.write(f"> **Report Date**: `{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}`  \n")
        rf.write(f"> **Audit Speed**: ` {total_topics} topics in {duration:.1f} seconds ` ({total_topics/duration:.1f} topics/sec)  \n")
        rf.write(f"> **Stage Source File**: [{os.path.basename(STAGE_MD_PATH)}](file://{STAGE_MD_PATH})  \n")
        rf.write(f"> **Production Source File**: [{os.path.basename(PROD_MD_PATH)}](file://{PROD_MD_PATH})  \n\n")
        
        rf.write(f"## 📈 Content Parity Score\n\n")
        rf.write(f"### **Content Match Percentage**: ` {match_pct:.2f}% `\n\n")
        
        rf.write(f"| Parity Metric | Stage Environment | Production Environment | Status |\n")
        rf.write(f"| :--- | :---: | :---: | :---: |\n")
        rf.write(f"| **Total Character Length (Excl. Spaces)** | {stage_char_count} | {prod_char_count} | {'✅ Match' if stage_char_count == prod_char_count else '❌ Drifted'} |\n")
        rf.write(f"| **Word Count** | {stage_words} | {prod_words} | {'✅ Match' if stage_words == prod_words else '❌ Drifted'} |\n")
        rf.write(f"| **Bold Text Blocks (`**`)** | {stage_bold_cnt} | {prod_bold_cnt} | {'✅ Match' if stage_bold_cnt == prod_bold_cnt else '❌ Drifted'} |\n")
        rf.write(f"| **Italic Text Blocks (`*`)** | {stage_italic_cnt} | {prod_italic_cnt} | {'✅ Match' if stage_italic_cnt == prod_italic_cnt else '❌ Drifted'} |\n")
        rf.write(f"| **Related / Inline Links** | {stage_links_cnt} | {prod_links_cnt} | {'✅ Match' if stage_links_cnt == prod_links_cnt else '❌ Drifted'} |\n\n")
        
        rf.write(f"## 🔗 Inline Links Parity Audit\n\n")
        rf.write(f"* **Stage Links**: `{len(stage_link_set)}`\n")
        rf.write(f"* **Production Links**: `{len(prod_link_set)}`\n\n")
        
        if missing_links_in_prod:
            rf.write(f"### ❌ Links in Stage but MISSING in Production ({len(missing_links_in_prod)})\n\n")
            for url in missing_links_in_prod:
                rf.write(f"- Label: `{stage_link_map.get(url, 'N/A')}`  \n  URL: `{url}`\n")
        else:
            rf.write(f"🎉 **No Stage links are missing in Production.**\n\n")
            
        if missing_links_in_stage:
            rf.write(f"### ➕ Extra Links in Production (Missing in Stage) ({len(missing_links_in_stage)})\n\n")
            for url in missing_links_in_stage:
                rf.write(f"- Label: `{prod_link_map.get(url, 'N/A')}`  \n  URL: `{url}`\n")
        else:
            rf.write(f"🎉 **No Production-only links found.**\n\n")

        rf.write(f"## 🔠 Per-Topic Missing Bold Text in Stage\n\n")
        if per_topic_bold_drift:
            rf.write(f"The following topics have **bold/strong** styled text in Production that is completely missing or unstyled in Stage:\n\n")
            for drift in per_topic_bold_drift:
                rf.write(f"### 📄 {drift['title']}\n")
                rf.write(f"- **Production URL**: [{drift['url']}]({drift['url']})\n")
                rf.write(f"- **Missing Bold Phrases in Stage**:\n")
                for pb in drift['missing_bolds']:
                    rf.write(f"  - [ ] `{pb}`\n")
                rf.write(f"\n")
        else:
            rf.write(f"🎉 **Perfect! No topics have bold text in Production that is missing in Stage.**\n\n")

        rf.write(f"## 📄 Per-Topic Sentence Parity Drift\n\n")
        if per_topic_sentence_drift:
            rf.write(f"The following topics have sentence-level mismatches between environments:\n\n")
            for drift in per_topic_sentence_drift:
                rf.write(f"### 📄 {drift['title']}\n")
                rf.write(f"- **Production URL**: [{drift['url']}]({drift['url']})\n")
                if drift['missing_in_prod']:
                    rf.write(f"- **Sentences in Stage but MISSING in Production**:\n")
                    for sent in drift['missing_in_prod']:
                        rf.write(f"  - \"{sent}\"\n")
                if drift['missing_in_stage']:
                    rf.write(f"- **Sentences in Production but MISSING in Stage**:\n")
                    for sent in drift['missing_in_stage']:
                        rf.write(f"  - \"{sent}\"\n")
                rf.write(f"\n")
        else:
            rf.write(f"🎉 **Perfect! No topics have drifted or missing sentences.**\n\n")

        rf.write(f"## 🔠 Bold Text Content Parity Summary\n\n")
        rf.write(f"* **Total Unique Bold Phrases Extracted**: `{total_unique_bolds}`\n")
        rf.write(f"* **Perfect Matches**: `{len(matched_bolds)}`\n")
        rf.write(f"* **Missing in Production**: `{len(missing_in_prod)}` (Stage bold phrases not styled as bold in Prod)\n")
        rf.write(f"* **Extra in Production**: `{len(extra_in_prod)}` (Prod bold phrases not styled as bold in Stage)\n")
        rf.write(f"* **Bold Text Parity Score**: ` {bold_match_pct:.2f}% `\n\n")
        rf.write(f"👉 **For the full side-by-side Bold Parity Checklist Report, see**: [bold_text_parity_report.md](file://{BOLD_REPORT_PATH})  \n\n")

        if diff_lines:
            rf.write(f"## 🔍 Parity Diff Highlights (First 100 Diff Lines)\n\n")
            rf.write(f"```diff\n")
            for line in diff_lines[:100]:
                rf.write(line + "\n")
            if len(diff_lines) > 100:
                rf.write(f"\n... (truncated {len(diff_lines) - 100} more difference lines)\n")
            rf.write(f"```\n")
        else:
            rf.write(f"🎉 **Perfect Match! The consolidated markdown files are identical.**\n")

    # ── Excel Parity Report Generation ──────────────────────────────────
    excel_filename = "N/A"
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        UI_REPORTS_DIR = os.path.join(os.getcwd(), '.ui_reports')
        os.makedirs(UI_REPORTS_DIR, exist_ok=True)
        
        # Dynamically follow REPORT_FILENAME if provided in environment, otherwise create clean deep-content-validation-<timestamp>.xlsx
        excel_filename = os.environ.get('REPORT_FILENAME') or f'deep-content-validation-{int(time.time())}.xlsx'
        excel_path = os.path.join(UI_REPORTS_DIR, excel_filename)
        
        print(f"📊 Exporting detailed Excel Parity Report to: {excel_path}...")
        
        # 1. Summary DataFrame
        summary_rows = [
            ['Avaya Content Parity & Verification Dashboard - Content validation', ''],
            ['Date Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Validation Speed', f"{total_topics} topics in {duration:.1f} seconds"],
            ['Content Parity Match Percentage', f"{match_pct:.2f}%"],
            ['', ''],
            ['📊 PARITY STATISTICS', '', '', '', ''],
            ['Metric Description', 'Stage', 'Production', 'Variance', 'Status'],
            ['Character Count (Excl. Spaces)', stage_char_count, prod_char_count, prod_char_count - stage_char_count, 'Match' if stage_char_count == prod_char_count else 'Drifted'],
            ['Word Count', stage_words, prod_words, prod_words - stage_words, 'Match' if stage_words == prod_words else 'Drifted'],
            ['Bold Phrase Blocks', stage_bold_cnt, prod_bold_cnt, prod_bold_cnt - stage_bold_cnt, 'Match' if stage_bold_cnt == prod_bold_cnt else 'Drifted'],
            ['Italic Phrase Blocks', stage_italic_cnt, prod_italic_cnt, prod_italic_cnt - stage_italic_cnt, 'Match' if stage_italic_cnt == prod_italic_cnt else 'Drifted'],
            ['Inline Links Count', stage_links_cnt, prod_links_cnt, prod_links_cnt - stage_links_cnt, 'Match' if stage_links_cnt == prod_links_cnt else 'Drifted'],
        ]
        summary_df = pd.DataFrame(summary_rows)
        
        # 2. Bold Parity DataFrame
        bold_rows = []
        for drift in per_topic_bold_drift:
            for pb in drift.get('missing_bolds', []):
                bold_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Bold text present in Production but MISSING in Stage',
                    'Bold Text / Phrase': pb
                })
            for pb in drift.get('extra_bolds', []):
                bold_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Bold text present in Stage but MISSING in Production',
                    'Bold Text / Phrase': pb
                })
        bold_df = pd.DataFrame(bold_rows) if bold_rows else pd.DataFrame(columns=['Topic Title', 'Production URL', 'Drift Status', 'Bold Text / Phrase'])
        
        # 3. Sentence Parity DataFrame
        sent_rows = []
        for drift in per_topic_sentence_drift:
            for sent in drift.get('missing_in_prod', []):
                sent_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Sentence present in Stage but MISSING in Production',
                    'Sentence Text': sent
                })
            for sent in drift.get('missing_in_stage', []):
                sent_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Sentence present in Production but MISSING in Stage',
                    'Sentence Text': sent
                })
        sent_df = pd.DataFrame(sent_rows) if sent_rows else pd.DataFrame(columns=['Topic Title', 'Production URL', 'Drift Status', 'Sentence Text'])
        
        # 4. Links Parity DataFrame
        link_rows = []
        for drift in per_topic_link_drift:
            for url in drift.get('missing_in_prod', []):
                link_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Link present in Stage but MISSING in Production',
                    'Link Label Text': drift['stage_map'].get(url, 'N/A'),
                    'Link Target URL': url
                })
            for url in drift.get('missing_in_stage', []):
                link_rows.append({
                    'Topic Title': drift['title'],
                    'Production URL': drift['url'],
                    'Drift Status': 'Link present in Production but MISSING in Stage',
                    'Link Label Text': drift['prod_map'].get(url, 'N/A'),
                    'Link Target URL': url
                })
        link_df = pd.DataFrame(link_rows) if link_rows else pd.DataFrame(columns=['Topic Title', 'Production URL', 'Drift Status', 'Link Label Text', 'Link Target URL'])
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            summary_df.to_excel(writer, sheet_name='Summary', header=False, index=False)
            bold_df.to_excel(writer, sheet_name='Missing & Extra Bold Text', index=False)
            sent_df.to_excel(writer, sheet_name='Missing & Extra Sentences', index=False)
            link_df.to_excel(writer, sheet_name='Missing & Extra Links', index=False)
            
            # Styling colors
            avaya_red_fill = PatternFill(start_color='DA291C', end_color='DA291C', fill_type='solid')
            zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
            white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
            header_font = Font(name='Segoe UI', size=11, color='FFFFFF', bold=True)
            bold_font = Font(name='Segoe UI', size=11, bold=True)
            regular_font = Font(name='Segoe UI', size=11)
            title_font = Font(name='Segoe UI', size=16, color='DA291C', bold=True)
            
            thin_border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            
            # Style Summary Sheet
            ws_summary = writer.sheets['Summary']
            ws_summary.merge_cells('A1:E1')
            ws_summary['A1'].font = title_font
            ws_summary.row_dimensions[1].height = 35
            ws_summary['A1'].alignment = Alignment(vertical='center')
            
            # Style Statistics Headers
            for r_idx in range(6, 13):
                ws_summary.row_dimensions[r_idx].height = 24
                for col_idx in range(1, 6):
                    cell = ws_summary.cell(row=r_idx, column=col_idx)
                    cell.font = regular_font
                    if r_idx == 6:
                        cell.fill = avaya_red_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    elif r_idx == 7:
                        cell.font = bold_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = thin_border
                    else:
                        cell.border = thin_border
                        if col_idx > 1:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Style Data Sheets
            for name in ['Missing & Extra Bold Text', 'Missing & Extra Sentences', 'Missing & Extra Links']:
                ws = writer.sheets[name]
                ws.row_dimensions[1].height = 28
                
                # Style Header Row
                for cell in ws[1]:
                    cell.fill = avaya_red_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Style Body Rows
                for row_idx in range(2, ws.max_row + 1):
                    ws.row_dimensions[row_idx].height = 20
                    is_even = (row_idx % 2 == 0)
                    row_fill = zebra_fill if is_even else white_fill
                    
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = row_fill
                        cell.font = regular_font
                        cell.border = thin_border
                        if col_idx == 3:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            # Color coding the Status column dynamically
                            if 'MISSING' in str(cell.value or ''):
                                cell.font = Font(name='Segoe UI', size=11, color='DA291C', bold=True)
                            else:
                                cell.font = Font(name='Segoe UI', size=11, color='1D4ED8', bold=True)
                
                # Auto-fit Column Widths beautifully
                for col in ws.columns:
                    max_len = 0
                    for cell in col:
                        if cell.row == 1:
                            continue  # Skip header for fitting to allow clean layout
                        max_len = max(max_len, len(str(cell.value or '')))
                    col_letter = col[0].column_letter
                    # Use a sensible cap of 70 characters width to keep the layout extremely clean
                    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 70)
        
        print(f"✅ Excel report written successfully to {excel_path}")
    except Exception as e:
        print(f"⚠️ Warning: Failed to generate Excel report. Error: {e}")
        import traceback
        traceback.print_exc()

    print(f"\n📊 Match Percentage: {match_pct:.2f}%")
    print(f"✓ Summary report saved: {COMPARISON_REPORT_PATH}")
    
    # standard format outputs
    print(f"::MATCH_PERCENTAGE::{match_pct:.2f}%")
    print(f"::STAGE_WORDS::{stage_words}")
    print(f"::PROD_WORDS::{prod_words}")
    print(f"::STAGE_BOLD::{stage_bold_cnt}")
    print(f"::PROD_BOLD::{prod_bold_cnt}")
    print(f"::STAGE_ITALIC::{stage_italic_cnt}")
    print(f"::PROD_ITALIC::{prod_italic_cnt}")
    print(f"::STAGE_LINKS::{stage_links_cnt}")
    print(f"::PROD_LINKS::{prod_links_cnt}")
    
    # Structured JSON for backend server integration
    import json
    results_json = {
        "match_percentage": f"{match_pct:.2f}%",
        "stage_words": stage_words,
        "prod_words": prod_words,
        "stage_bold": stage_bold_cnt,
        "prod_bold": prod_bold_cnt,
        "stage_italic": stage_italic_cnt,
        "prod_italic": prod_italic_cnt,
        "stage_links": stage_links_cnt,
        "prod_links": prod_links_cnt,
        "stage_char": stage_char_count,
        "prod_char": prod_char_count,
        "stage_topics": total_topics,
        "prod_topics": total_topics,
        "duration": duration,
        "stage_md_file": "stage_topics.md",
        "prod_md_file": "prod_topics.md",
        "comparison_report": "md_comparison_report.md",
        "excel_report": excel_filename
    }
    print(f"::RESULTS::{json.dumps(results_json)}")

async def main():
    try:
        total, duration, prod_toc, stage_md_blocks, prod_md_blocks = await process_topics_in_markdown()
        compare_compiled_markdowns(total, duration, prod_toc, stage_md_blocks, prod_md_blocks)
        return 0
    except Exception as e:
        print(f"❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    asyncio.run(main())
