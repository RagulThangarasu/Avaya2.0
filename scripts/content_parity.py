import os
import sys
import json
import asyncio
import re
from datetime import datetime
from urllib.parse import urljoin, urlparse
import pandas as pd
from playwright.async_api import async_playwright

# Configurations
REPORTS_DIR = os.path.join(os.getcwd(), 'reports')
UI_REPORTS_DIR = os.path.join(os.getcwd(), '.ui_reports')
os.makedirs(REPORTS_DIR, exist_ok=True)
os.makedirs(UI_REPORTS_DIR, exist_ok=True)

TEST_URLS_PATH = os.path.join(os.getcwd(), 'config', 'test-urls.json')
AUTH_STATE_PATH = os.path.join(os.getcwd(), 'auth-sessions', 'storage-state.json')

STAGE_URL = os.environ.get('STAGE_URL')
PROD_URL = os.environ.get('PROD_URL')
REPORT_FILENAME = os.environ.get('REPORT_FILENAME') or os.path.join(UI_REPORTS_DIR, f'content-parity-{int(datetime.now().timestamp())}.xlsx')

if not STAGE_URL or not PROD_URL:
    try:
        with open(TEST_URLS_PATH, 'r') as f:
            config = json.load(f)
            STAGE_URL = STAGE_URL or config.get('stage')
            PROD_URL = PROD_URL or config.get('production')
    except:
        pass

if not STAGE_URL or not PROD_URL:
    print("❌ Error: Missing STAGE_URL or PROD_URL.")
    sys.exit(1)

print(f"🚀 Starting TOC Validation")
print(f"   Stage (Publish): {STAGE_URL}")
print(f"   Production:      {PROD_URL}")

def slugify(t):
    if not t: return ""
    return re.sub(r'[^a-z0-9]', '', t.lower())

def get_filename(url):
    path = urlparse(url).path
    filename = path.split('/')[-1] if '/' in path else path
    return filename.replace('.html', '').replace('.htm', '').lower().replace('-', '').replace('_', '')

async def handle_cookies(page):
    try:
        for sel in ['#onetrust-accept-btn-handler', '#btn-accept-all', 'button:has-text("Accept")', '.cookie-accept']:
            if await page.locator(sel).is_visible(timeout=1500):
                await page.click(sel)
                await page.wait_for_timeout(500)
                break
    except:
        pass

# ── Prod TOC Extraction ──────────────────────────────────────────────
async def extract_prod_toc(page, base_url):
    await page.wait_for_load_state("networkidle")
    await handle_cookies(page)
    await page.wait_for_timeout(3000)

    try:
        # 1. Click "Expand All" button if available
        expand_btn = page.locator('.zDocsCollapseExpandButton').first
        if await expand_btn.is_visible(timeout=5000):
            await expand_btn.click()
            await page.wait_for_timeout(5000)
    except:
        pass

    # 2. Iteratively expand nested nodes up to 5 levels deep to capture full sequence
    for level in range(5):
        try:
            # Trigger parallel clicks in-browser to bypass Playwright's sequential mouse emulation (100x faster!)
            expanded_count = await page.evaluate('''() => {
                const buttons = Array.from(document.querySelectorAll('.zDocsTocCollapseItemButton[aria-expanded="false"], button[aria-expanded="false"], .expand-icon'));
                buttons.forEach(btn => {
                    try { btn.click(); } catch(e) {}
                });
                return buttons.length;
            }''')
            if expanded_count == 0:
                break
            print(f"   [Level {level+1}] Triggered expansion on {expanded_count} nested nodes...")
            await page.wait_for_timeout(1000)
        except Exception as e:
            print(f"   [Level {level+1}] Expand warning: {e}")
            break

    links_data = await page.evaluate(r'''() => {
        const results = [];
        const container = document.querySelector('.zDocsTocList') || document.querySelector('.zDocsTOC');
        let allLinks = [];
        if (container) {
            allLinks = Array.from(container.querySelectorAll('a[href]'));
        }
        if (allLinks.length === 0) {
            allLinks = Array.from(document.querySelectorAll('nav a[href], [class*="sidebar"] a[href]'));
        }
        allLinks.forEach(a => {
            const href = a.getAttribute('href');
            let text = a.innerText.trim();
            // Clean up titles (remove leading/trailing symbols, newlines, tabs, chevron chars)
            text = text.replace(/[\r\n\t]+/g, ' ').replace(/^\s*[\u203A\u25BC\u25B6>v-]\s*/g, '').trim();
            if (href && text && !href.startsWith('#') && !href.startsWith('javascript')) {
                results.push({text, href});
            }
        });
        return results;
    }''')

    toc = []
    seen = set()
    for item in links_data:
        full_url = urljoin(base_url, item['href']).split('#')[0].split('?')[0]
        if full_url not in seen:
            toc.append({'title': item['text'], 'url': full_url})
            seen.add(full_url)

    print(f"   ✅ Prod TOC: {len(toc)} topics found.")
    return toc

# ── Stage/Published TOC Extraction ───────────────────────────────────
async def extract_stage_toc(page, base_url):
    await page.wait_for_load_state("networkidle")
    await handle_cookies(page)
    await page.wait_for_timeout(2000)

    # Iteratively expand collapsible nested items on Stage to match sequences
    for level in range(3):
        try:
            expanded_count = await page.evaluate('''() => {
                const buttons = Array.from(document.querySelectorAll('.cmp-navigation__item--active[aria-expanded="false"], .cmp-navigation__item[aria-expanded="false"], button[aria-expanded="false"]'));
                buttons.forEach(btn => {
                    try { btn.click(); } catch(e) {}
                });
                return buttons.length;
            }''')
            if expanded_count == 0:
                break
            await page.wait_for_timeout(500)
        except:
            break

    links_data = await page.evaluate(r'''() => {
        const results = [];
        const links = document.querySelectorAll('.cmp-navigation__item-link');
        links.forEach(a => {
            const href = a.getAttribute('href');
            let text = a.innerText.trim();
            text = text.replace(/[\r\n\t]+/g, ' ').replace(/^\s*[\u203A\u25BC\u25B6>v-]\s*/g, '').trim();
            if (href && text && !href.startsWith('#') && !href.startsWith('javascript')) {
                results.push({text, href});
            }
        });
        return results;
    }''')

    toc = []
    seen = set()
    for item in links_data:
        full_url = urljoin(base_url, item['href']).split('#')[0].split('?')[0]
        if full_url not in seen and '.html' in full_url:
            toc.append({'title': item['text'], 'url': full_url})
            seen.add(full_url)

    print(f"   ✅ Stage TOC: {len(toc)} topics found.")
    return toc

# ── Main Validation ──────────────────────────────────────────────────
async def run_validation():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)

        stage_ctx = await browser.new_context(
            storage_state=AUTH_STATE_PATH if os.path.exists(AUTH_STATE_PATH) else None
        )
        prod_ctx = await browser.new_context()

        print("\n🔍 Extracting TOC from both environments...")

        # Extract Prod and Stage TOC concurrently in parallel (saves 50% load duration!)
        async def crawl_prod():
            p_page = await prod_ctx.new_page()
            await p_page.goto(PROD_URL, wait_until="networkidle", timeout=60000)
            toc = await extract_prod_toc(p_page, PROD_URL)
            await p_page.close()
            return toc

        async def crawl_stage():
            s_page = await stage_ctx.new_page()
            await s_page.goto(STAGE_URL, wait_until="networkidle", timeout=60000)
            toc = await extract_stage_toc(s_page, STAGE_URL)
            await s_page.close()
            return toc

        prod_toc, stage_toc = await asyncio.gather(crawl_prod(), crawl_stage())

        print(f"\n📊 Comparing: Prod={len(prod_toc)} vs Stage={len(stage_toc)}")

        # ── Build lookup maps ────────────────────────────────────────
        prod_fn_map = {}  # filename -> (index, item)
        for i, t in enumerate(prod_toc):
            fn = get_filename(t['url'])
            if fn:
                prod_fn_map[fn] = (i, t)

        stage_fn_map = {} # filename -> list of (index, item, used)
        for i, t in enumerate(stage_toc):
            fn = get_filename(t['url'])
            if fn:
                if fn not in stage_fn_map:
                    stage_fn_map[fn] = []
                stage_fn_map[fn].append({'idx': i, 'item': t, 'used': False})

        # ── Detailed Comparison ──────────────────────────────────────
        comparison = []
        used_stage = set()

        for prod_idx, prod_item in enumerate(prod_toc):
            prod_fn = get_filename(prod_item['url'])
            
            # Match by filename
            stage_idx = None
            stage_item = None
            
            if prod_fn in stage_fn_map:
                candidates = stage_fn_map[prod_fn]
                for cand in candidates:
                    if not cand['used']:
                        stage_idx = cand['idx']
                        stage_item = cand['item']
                        cand['used'] = True
                        break

            if stage_item:
                # Found in both - check details
                title_match = slugify(prod_item['title']) == slugify(stage_item['title'])
                same_position = (prod_idx == stage_idx)
                
                if title_match and same_position:
                    status = '✅ MATCH'
                    issue = ''
                else:
                    status = '❌ MISMATCH'
                    reasons = []
                    if not title_match: reasons.append('Title')
                    if not same_position: reasons.append('Order')
                    issue = f"Differences in: {', '.join(reasons)}"
                
                comparison.append({
                    'Prod Sequence': prod_idx + 1,
                    'Stage Sequence': stage_idx + 1,
                    'Sequence Match': '✅' if same_position else '❌',
                    'Content Match': '✅' if title_match else '❌',
                    'Prod Title': prod_item['title'],
                    'Stage Title': stage_item['title'],
                    'Prod URL': prod_item['url'],
                    'Stage URL': stage_item['url'],
                })
            else:
                # Missing in stage
                comparison.append({
                    'Prod Sequence': prod_idx + 1,
                    'Stage Sequence': '-',
                    'Sequence Match': '❌',
                    'Content Match': '❌',
                    'Prod Title': prod_item['title'],
                    'Stage Title': '[MISSING]',
                    'Prod URL': prod_item['url'],
                    'Stage URL': '-',
                })

        # Find items only in Stage (extra)
        for fn, candidates in stage_fn_map.items():
            for cand in candidates:
                if not cand['used']:
                    stage_idx = cand['idx']
                    stage_item = cand['item']
                    comparison.append({
                    'Prod Sequence': '-',
                    'Stage Sequence': stage_idx + 1,
                    'Sequence Match': '❌',
                    'Content Match': '❌',
                    'Prod Title': '[MISSING]',
                    'Stage Title': stage_item['title'],
                    'Prod URL': '-',
                    'Stage URL': stage_item['url'],
                })

        # ── Build structure data for stats ──────────────────────────
        max_len = max(len(stage_toc), len(prod_toc))
        structure_data = []
        for i in range(max_len):
            s_item = stage_toc[i] if i < len(stage_toc) else {'title': '', 'url': ''}
            p_item = prod_toc[i] if i < len(prod_toc) else {'title': '', 'url': ''}
            # Determine if there's a mismatch at this position
            is_mismatch = (s_item['title'] != p_item['title']) if (s_item['title'] and p_item['title']) else False
            if not s_item['title'] or not p_item['title']:
                is_mismatch = True  # Missing in one environment
            structure_data.append({
                'Stage #': i + 1 if i < len(stage_toc) else '',
                'Stage Title': s_item['title'],
                'Prod #': i + 1 if i < len(prod_toc) else '',
                'Prod Title': p_item['title'],
                'Status': '❌ MISMATCH' if is_mismatch else '✅ MATCH'
            })

        # ── Calculate Stats from Structure Data ──────────────────────
        total = len(structure_data)
        full_match = len([r for r in structure_data if r['Status'] == '✅ MATCH'])
        mismatch = len([r for r in structure_data if r['Status'] == '❌ MISMATCH'])
        match_pct = int(full_match / max(total, 1) * 100)

        print(f"\n{'='*60}")
        print(f"  📈 TOC Validation Results")
        print(f"{'='*60}")
        print(f"  ✅ Match:       {full_match}")
        print(f"  ❌ Mismatch:    {mismatch}")
        print(f"  📊 Match Rate:  {match_pct}%")
        print(f"{'='*60}\n")

        # ── Generate Excel Report ────────────────────────────────────
        os.makedirs(os.path.dirname(REPORT_FILENAME), exist_ok=True)

        # Prepare individual TOC dataframes
        stage_toc_df = pd.DataFrame([
            {'#': i + 1, 'Title': t['title'], 'URL': t['url']} 
            for i, t in enumerate(stage_toc)
        ])
        prod_toc_df = pd.DataFrame([
            {'#': i + 1, 'Title': t['title'], 'URL': t['url']} 
            for i, t in enumerate(prod_toc)
        ])
        
        # Remove comparison variable as we're now using structure_data for stats
        comparison = None
        summary_data = [
            ['TOC Validation Report'],
            [''],
            ['Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Stage URL', STAGE_URL],
            ['Prod URL', PROD_URL],
            [''],
            ['── Environment Statistics ──'],
            ['Prod Topics', len(prod_toc)],
            ['Stage Topics', len(stage_toc)],
            [''],
            ['── Validation Results ──'],
            ['✅ Match', full_match],
            ['❌ Mismatch', mismatch],
            ['Total Compared', total],
            [''],
            ['── Match Rate ──'],
            ['Percentage', f'{match_pct}%'],
        ]

        with pd.ExcelWriter(REPORT_FILENAME, engine='openpyxl') as writer:
            # 1. Summary
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', header=False, index=False)
            
            # 2. Stage TOC
            stage_toc_df.to_excel(writer, sheet_name='Stage TOC', index=False)
            
            # 3. Prod TOC
            prod_toc_df.to_excel(writer, sheet_name='Prod TOC', index=False)

            # 4. Structure (Side-by-side comparison with status)
            structure_df = pd.DataFrame(structure_data)
            structure_df.to_excel(writer, sheet_name='Structure', index=False)
            
            # Apply red highlighting to mismatches in Structure sheet
            from openpyxl.styles import PatternFill
            ws = writer.sheets['Structure']
            red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
            
            for row_idx, row_data in enumerate(structure_data, start=2):  # Start at 2 (after header)
                if row_data['Status'] == '❌ MISMATCH':
                    for col_idx in range(1, 6):  # Columns A-E
                        ws.cell(row=row_idx, column=col_idx).fill = red_fill

        results = {
            'overall': match_pct,
            'matched': full_match,
            'mismatch': mismatch,
        }
        print(f"\n✅ Report saved: {REPORT_FILENAME}")
        print(f"::RESULTS::{json.dumps(results)}")
        
        await browser.close()

if __name__ == "__main__":
    asyncio.run(run_validation())
