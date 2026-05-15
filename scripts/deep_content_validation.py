"""
deep_content_validation.py - Full Content Extraction & Validation
Extracts ALL content from each topic page, validates, and generates TXT + XLSX reports
"""
import os, sys, json, asyncio, re
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
from playwright.async_api import async_playwright

UI_REPORTS_DIR  = os.path.join(os.getcwd(), '.ui_reports')
os.makedirs(UI_REPORTS_DIR, exist_ok=True)

AUTH_STATE_PATH = os.path.join(os.getcwd(), 'auth-sessions', 'storage-state.json')
TEST_URLS_PATH  = os.path.join(os.getcwd(), 'config', 'test-urls.json')
CONCURRENCY     = 1


def sim(a, b):
    if not a and not b: return 100
    if not a or not b:  return 0
    return int(SequenceMatcher(None, a.lower().strip(), b.lower().strip()).ratio() * 100)


async def handle_cookies(page):
    """Aggressively handle cookie banners and login popups that block content"""
    try:
        # Common selectors for Accept buttons and Close buttons on modals
        await page.evaluate("""() => {
            const selectors = [
                '#onetrust-accept-btn-handler', '#btn-accept-all', 
                'button[aria-label*="Accept"]', '.cookie-accept',
                '.osano-cm-accept-all', '#accept-recommended-btn-handler',
                '.login-popup-close', '[aria-label="Close login popup"]',
                '.modal-close', '.close-button', '.zDocsLoginModal .close'
            ];
            selectors.forEach(s => {
                try {
                    const el = document.querySelector(s);
                    if (el && el.offsetParent !== null) { // if visible
                        el.click();
                        console.log('Clicked: ' + s);
                    }
                } catch(e) {}
            });
            
            // Specifically for the "Log in to get a better experience" popup
            const loginPopup = document.querySelector('.login-prompt, [class*="LoginModal"]');
            if (loginPopup) {
                const closeBtn = loginPopup.querySelector('button, .close, [aria-label*="Close"]');
                if (closeBtn) closeBtn.click();
            }

            // Find buttons by text (e.g., "Accept All Cookies")
            const allButtons = Array.from(document.querySelectorAll('button, a'));
            const acceptBtn = allButtons.find(b => {
                const txt = (b.innerText || '').trim();
                return txt === 'Accept All Cookies' || txt === 'Accept All' || txt === 'I Accept';
            });
            if (acceptBtn && acceptBtn.offsetParent !== null) {
                acceptBtn.click();
                console.log('Clicked button by text: ' + acceptBtn.innerText);
            }

            // Target the "Avaya is refreshing" banner specifically
            const banners = [...document.querySelectorAll('div, section, aside')];
            const splash = banners.find(b => b.innerText && b.innerText.includes('Avaya is refreshing'));
            if (splash) {
                // Try to find a close button within it
                const xBtn = splash.querySelector('button, .close, [class*="close"], [aria-label*="close"]');
                if (xBtn) {
                    xBtn.click();
                } else {
                    // Fallback: just remove it from DOM
                    splash.style.display = 'none';
                    splash.remove();
                }
            }
        }""")
        await asyncio.sleep(1)
    except:
        pass


async def extract_toc(page, base_url, is_stage=False):
    env = 'Stage' if is_stage else 'Prod'
    print(f"  📚 Extracting {env} TOC...")
    try:
        wait_until = 'networkidle' if not is_stage else 'domcontentloaded'
        await page.goto(base_url, wait_until=wait_until, timeout=90000)
        await asyncio.sleep(3)
        await handle_cookies(page)
        
        # Check for splash screen "Avaya is refreshing"
        for _ in range(5):
            is_refreshing = await page.evaluate("() => document.body.innerText.includes('Avaya is refreshing')")
            if not is_refreshing: break
            print(f"  ⏳ Waiting for splash screen to disappear on {env}...")
            await asyncio.sleep(2)
            await handle_cookies(page)
    except Exception as e:
        print(f"  ❌ Failed: {e}"); return []

    bundle_m = re.search(r'/bundle/([^/]+)', base_url)
    bundle   = bundle_m.group(1) if bundle_m else ''

    if not is_stage:
        print(f"  🔄 Expanding all TOC nodes...")
        for rnd in range(30):
            n = await page.evaluate('''() => {
                const btn = document.querySelector('.zDocsCollapseExpandButton');
                if (btn) btn.click();
                const toggles = [...document.querySelectorAll('.zDocsTocItemCollapsed .zDocsTocItemToggle[aria-expanded="false"]')];
                toggles.forEach(t => { try { t.click(); } catch(e){} });
                return toggles.length;
            }''')
            await asyncio.sleep(0.5)
            if n == 0 and rnd > 2: break
        await page.evaluate('''async () => {
            const toc = document.querySelector('.zDocsTocList') || document.documentElement;
            for (let i = 0; i < 100; i++) { toc.scrollTop += 250; await new Promise(r => setTimeout(r, 50)); }
        }''')

    links = await page.evaluate(f'''() => {{
        const bundle = {json.dumps(bundle)};
        const isStage = {json.dumps(is_stage)};
        const seen = new Set(), results = [];
        
        // Target specific TOC containers to avoid header/footer links
        const sels = isStage 
            ? ['nav.cmp-navigation', '.cmp-navigation', '.left-nav', 'aside nav'] 
            : ['.zDocsTocList', '.zDocsTOC', '.toc-container', 'nav'];
            
        let container = null;
        for (const s of sels) {{
            try {{ const el = document.querySelector(s); if (el && el.querySelectorAll('a[href]').length > 3) {{ container = el; break; }} }} catch(e){{}}
        }}
        container = container || document.body;
        
        container.querySelectorAll('a[href]').forEach(a => {{
            try {{
                const raw = a.getAttribute('href') || '';
                if (!raw || raw.startsWith('#') || raw.startsWith('javascript')) return;
                
                const urlObj = new URL(raw, window.location.href);
                const url = urlObj.href.split('#')[0].split('?')[0];
                const text = (a.innerText || a.textContent || '').trim().replace(/\\s+/g,' ');
                
                if (!text || text.length < 3 || text.length > 250) return;
                
                // VALIDATION: Must be a topic page, not a bundle root or category
                const isTopic = url.includes('/page/') || url.endsWith('.html');
                if (!isTopic) return;
                
                // Ensure it belongs to the current bundle
                if (bundle && !url.includes(bundle)) return;
                
                // AVOID BUNDLE ROOT: If URL ends with bundle name and doesn't have /page/
                const isBundleRoot = url.split('/').pop() === bundle || url.split('/').pop() === '';
                if (isBundleRoot && !url.includes('/page/')) return;

                if (seen.has(url)) return;
                seen.add(url); 
                results.push({{ text, url }});
            }} catch(e) {{}}
        }});
        return results;
    }}''')

    NOISE = {'cookie','share','download','export','feedback','search','login','logout','next topic','previous topic'}
    return [{'title': l['text'], 'url': l['url']} for l in links
            if l.get('url') and l.get('text') and not any(n in l['text'].lower() for n in NOISE)]


async def get_full_content(page, url, is_stage=False):
    """Extract complete page content using env-specific selectors with aggressive waiting and scrolling"""
    try:
        # 1. Load and wait for content to be meaningful
        # Use networkidle for Production to ensure React components load
        wait_until = 'networkidle' if not is_stage else 'domcontentloaded'
        await page.goto(url, wait_until=wait_until, timeout=90000)
        await asyncio.sleep(2)
        await handle_cookies(page)        # Check for splash screen "Avaya is refreshing" or cookie banners
        for _ in range(10):
            is_blocked = await page.evaluate("""() => {
                const text = document.body.innerText;
                const splash = text.includes('Avaya is refreshing') || text.includes('refreshed Documentation Center');
                const loading = document.querySelector('.loading, .spinner, #loading-mask');
                const cookieBanner = document.querySelector('#onetrust-banner-sdk, .ot-sdk-container') || 
                                     [...document.querySelectorAll('button')].find(b => (b.innerText||'').includes('Accept All Cookies'));
                return splash || !!loading || (cookieBanner && cookieBanner.offsetParent !== null);
            }""")
            if not is_blocked: break
            print(f"  ⏳ Page seems blocked on {url[:80]}... handling cookies and waiting")
            await handle_cookies(page)
            await asyncio.sleep(2)
            # Try to force remove it if it's still there after a few tries
            if _ > 5:
                await page.evaluate("""() => {
                    const selectors = ['#onetrust-banner-sdk', '.ot-sdk-container', '.loading', '.spinner', '#loading-mask'];
                    selectors.forEach(s => { const el = document.querySelector(s); if (el) el.remove(); });
                }""")
        
        # Aggressive wait for content to appear (beyond networkidle)
        try:
            await page.wait_for_function("""() => {
                const sel = '.zDocsTopicPageBody, .zDocsTopicPageBodyContent, .zDocsTopicPage, .topicbody, article, .topic-renderer__content, main';
                const el = document.querySelector(sel);
                if (!el) return false;
                
                // Clone and remove action-bar noise before checking text length
                const clone = el.cloneNode(true);
                const noiseSelectors = [
                    '.zDocsTopicActionBar', '.zDocsTopicActions', '.zDocsActionBar',
                    '[class*="ActionBar"]', '[class*="action-bar"]',
                    '.zDocsAddToMyTopics', '.zDocsDownloadPdf', '.zDocsWatchTopic',
                    '[class*="AddToMyTopics"]', '[class*="DownloadPdf"]',
                    '.topic-actions', '.page-actions', '.action-menu',
                    '.zDocsTopicPageTocContainer', '.zDocsTopicPageHead'
                ];
                noiseSelectors.forEach(s => {
                    try { clone.querySelectorAll(s).forEach(n => n.remove()); } catch(e) {}
                });
                
                // We want to see some actual paragraphs or documentation text
                const text = clone.innerText.trim();
                return text.length > 50; // Lower threshold but specifically looking for content
            }""", timeout=20000)
        except:
            print(f"  ⚠️ Timeout waiting for meaningful text content on {url[:80]}")

        # 2. Scroll to trigger any lazy-loaded tables/images
        await page.evaluate("""async () => {
            for (let i = 0; i < 5; i++) { window.scrollBy(0, 800); await new Promise(r => setTimeout(r, 200)); }
            window.scrollTo(0, 0);
        }""")
        await asyncio.sleep(1)
        await handle_cookies(page) # Check again after scroll

        # 3. Extract content
        content = await page.evaluate(f"""(isStage) => {{
            // For PROD: use .zDocsTopicPageBody as the PRIMARY container (this is the main topic body on documentation.avaya.com)
            // For STAGE: use topic-renderer or cmp-topic-renderer
            let main = null;
            if (!isStage) {{
                // PROD priority order - zDocsTopicPageBody is the correct class
                main = document.querySelector('.zDocsTopicPageBody') ||
                       document.querySelector('.zDocsTopicPageBodyContent') ||
                       document.querySelector('.zDocsTopicPage') ||
                       document.querySelector('.topicbody') ||
                       document.querySelector('article');
            }} else {{
                main = document.querySelector('.topic-renderer__content') ||
                       document.querySelector('.cmp-topic-renderer') ||
                       document.querySelector('.topic-renderer') ||
                       document.querySelector('main') ||
                       document.querySelector('article');
            }}
            
            if (!main) {{
                main = document.querySelector('main, article, .content, #content');
            }}
            if (!main) {{
                const containers = [...document.querySelectorAll('div')].filter(d => d.innerText.length > 500);
                if (containers.length > 0) main = containers.sort((a, b) => b.innerText.length - a.innerText.length)[0];
            }}
            
            main = main || document.body;

            // DONT pick up header/footer/nav/action-bars — REMOVE them from DOM clone
            const clone = main.cloneNode(true);
            const exclude = [
                'header', 'footer', 'nav', 
                '.zDocsHeader', '.zDocsFooter', '.cmp-header', '.cmp-footer', 
                '.navigation', '.toc', '.zDocsTOC',
                // Action bar / toolbar UI elements (Add to My Topics, Download PDF, etc.)
                '.zDocsTopicActionBar', '.zDocsTopicActions', '.zDocsActionBar',
                '[class*="ActionBar"]', '[class*="action-bar"]',
                '.zDocsAddToMyTopics', '.zDocsDownloadPdf', '.zDocsWatchTopic',
                '[class*="AddToMyTopics"]', '[class*="DownloadPdf"]', '[class*="WatchTopic"]',
                '.topic-actions', '.page-actions', '.action-menu',
                // Breadcrumbs and metadata bars
                '.zDocsBreadcrumb', '[class*="breadcrumb"]', '[class*="Breadcrumb"]',
                // AI Summary containers
                '[class*="AiSummary"]', '[id*="AiSummary"]', '[class*="ai-summary"]',
                // Topic page head (title bar with action buttons)
                '.zDocsTopicPageHead', '.zDocsTopicPageTocContainer',
                // Share/Print/Download floating menus
                '[class*="ShareMenu"]', '[class*="PrintMenu"]', '[class*="DownloadMenu"]'
            ];
            exclude.forEach(s => {{
                try {{ 
                    clone.querySelectorAll(s).forEach(el => el.remove());
                }} catch(e) {{}}
            }});
            
            // Remove elements whose text is purely action-bar UI noise
            const actionNoisePatterns = [
                'Add to My Topics', 'Add Topic & Subtopics', 'Add Entire Publication',
                'Download PDF', 'Download selected topic', 'Watch', 'Share',
                'Add content to a collection by clicking'
            ];
            try {{
                clone.querySelectorAll('button, [role="button"], .zDocsTopicPageHead *').forEach(el => {{
                    const txt = (el.innerText || '').trim();
                    if (actionNoisePatterns.some(p => txt.includes(p))) el.remove();
                }});
                // Also remove any li/p that is purely about "Add to My Topics" UI instructions
                clone.querySelectorAll('li, p').forEach(el => {{
                    const txt = (el.innerText || '').trim();
                    if (txt.includes('Add to My Topics') && txt.includes('collection') && txt.length < 300) el.remove();
                    if (txt.includes('Add content to a collection')) el.remove();
                }});
            }} catch(e) {{}}

            const clean = (txt) => {{
                const t = (txt || '').trim().replace(/\\s+/g,' ');
                return t.length > 2 ? t : '';
            }};

            const title    = document.title || '';
            const h1       = [...clone.querySelectorAll('h1')].map(h => clean(h.innerText)).filter(Boolean);
            const h2       = [...clone.querySelectorAll('h2')].map(h => clean(h.innerText)).filter(Boolean);
            const h3       = [...clone.querySelectorAll('h3')].map(h => clean(h.innerText)).filter(Boolean);
            
            const bolds    = [...clone.querySelectorAll('b, strong')].map(b => clean(b.innerText)).filter(t => t.length > 2);
            const italics  = [...clone.querySelectorAll('i, em')].map(i => clean(i.innerText)).filter(t => t.length > 2);
            
            // Paragraphs
            const paras    = [...clone.querySelectorAll('p, div.zDocsTopicPara, .conbody p, .body p')]
                .map(p => clean(p.innerText)).filter(t => t.length > 5);
            
            // Lists: ul/ol structure + individual li items
            const ulElements = [...clone.querySelectorAll('ul, ol')];
            const ulCount    = ulElements.length;
            const lists      = [...clone.querySelectorAll('li, dd, dt')]
                .map(l => clean(l.innerText)).filter(t => t.length > 3);
            
            // Emphasized text (em tags specifically)
            const emTexts    = [...clone.querySelectorAll('em')].map(e => clean(e.innerText)).filter(Boolean);
            
            const fullText = (clone.innerText || '').replace(/\\s+/g,' ').trim();
            const bodyTextSnippet = fullText.slice(0, 1000);
            
            // Text breaking issue detection
            const textBreakingIssues = [];
            const containerWidth = clone.offsetWidth || 1280;
            
            // Check for very long words without spaces (often breaks layout)
            const longWords = fullText.split(/\\s+/).filter(w => w.length > 50);
            if (longWords.length > 5) textBreakingIssues.push('Very long words without breaks: ' + longWords.slice(0, 2).join(', '));
            
            // Check for encoding corruption (common in AEM→zDocs migration)
            const encodingIssues = [];
            const suspiciousPatterns = [
                /\\?\\?+/g,              // ???? repeated
                /[\\u0080-\\u00FF]{5,}/g, // Too many extended ASCII
                /&#?[0-9]{3,};/g,        // HTML entities
                /\\\\[uU][0-9a-fA-F]{4,8}/g  // Unicode escapes
            ];
            suspiciousPatterns.forEach(pattern => {{
                const matches = fullText.match(pattern);
                if (matches && matches.length > 3) {{
                    encodingIssues.push('Encoding corruption detected: ' + matches.slice(0, 2).join(', '));
                }}
            }});
            if (encodingIssues.length > 0) textBreakingIssues.push(...encodingIssues);
            
            // Check for text truncation (ends abruptly mid-word)
            if (fullText.length > 100) {{
                const lastChar = fullText[fullText.length - 1];
                const lastWord = fullText.split(/\\s+/).pop() || '';
                if (lastWord.length > 20 && !/[.!?,;:]/.test(lastChar)) {{
                    textBreakingIssues.push('Text appears truncated at: ' + lastWord.substring(0, 20) + '...');
                }}
            }}
            
            // Check for mixed character encodings
            const hasUnicodeSpaces = /[\\u00A0\\u2000-\\u200B\\u3000]/g.test(fullText);
            const hasMixedQuotes = /[""'']/g.test(fullText);
            if (hasUnicodeSpaces && hasMixedQuotes) {{
                textBreakingIssues.push('Mixed character encodings detected');
            }}
            
            // Check for line breaks preserved as visible text
            const lineBreaks = fullText.match(/\\n|\\r|<br>|<br\\s*\\/>|&lt;br&gt;/g);
            if (lineBreaks && lineBreaks.length > 3) {{
                textBreakingIssues.push('Formatting characters in text: ' + lineBreaks.length + ' line breaks detected');
            }}
            
            // Sentence count estimation
            const sentences = fullText.split(/[.!?]+/).filter(s => s.trim().length > 10);

            // Tables: extract row content for comparison + detect breaking issues
            const tableElements = [...clone.querySelectorAll('table, .cmp-table, [role="table"]')];
            const uniqueTables = tableElements.filter((el, idx) => {{
                return !tableElements.some((other, oIdx) => idx !== oIdx && other.contains(el));
            }});
            const tableBreakingIssues = [];
            const tableData = uniqueTables.filter(table => {{
                const rows = [...table.querySelectorAll('tr, [role="row"]')];
                const hasText = table.innerText.trim().length > 5;
                return rows.length > 0 && hasText;
            }}).map(table => {{
                const rows = [...table.querySelectorAll('tr, [role="row"]')];
                const headerRow = rows.length > 0 ? rows[0].innerText.trim().replace(/\\s+/g,' ') : '';
                const rowCount = rows.length;
                const cellCount = table.querySelectorAll('td, th').length;
                
                // Check for table breaking issues
                const tableWidth = table.offsetWidth || 0;
                const isOverflowing = tableWidth > containerWidth * 0.9;
                const hasLongCells = [...table.querySelectorAll('td, th')].some(cell => 
                    (cell.innerText || '').length > 100 && !cell.textContent.includes('\\n')
                );
                const hasNarrowColumns = [...table.querySelectorAll('tr')].some(row => 
                    [...row.querySelectorAll('td, th')].some(cell => cell.offsetWidth < 30)
                );
                
                if (isOverflowing) tableBreakingIssues.push('Table exceeds container width');
                if (hasNarrowColumns) tableBreakingIssues.push('Columns too narrow for content');
                if (hasLongCells && rowCount < 5) tableBreakingIssues.push('Long text not wrapping in cells');
                
                return {{ 
                    header: headerRow.substring(0, 200), rows: rowCount, cells: cellCount,
                    content: rows.slice(0, 5).map(tr => tr.innerText.trim().replace(/\\s+/g,' ').substring(0, 150)).join(' || '),
                    isOverflowing, hasNarrowColumns, hasLongCells,
                    breakingIssues: tableBreakingIssues.length > 0 ? tableBreakingIssues : null
                }};
            }});

            // Images: content images with breaking issue detection
            const imageBreakingIssues = [];
            const imageData = [...clone.querySelectorAll('img')].filter(img => {{
                const src = img.getAttribute('src') || '';
                if (!src || src.startsWith('data:')) return false;
                const filename = src.split('/').pop().split('?')[0].toLowerCase();
                // Always keep note/warning/caution/tip/important icons even if from clientlibs
                const isAdmonitionIcon = filename.includes('note') || filename.includes('warn') ||
                    filename.includes('caution') || filename.includes('tip') || filename.includes('important') ||
                    filename.includes('danger') || filename.includes('info') || filename.includes('alert');
                const parent = img.closest('.note,.warning,.caution,.tip,.important,.danger,.zDocsNote,.zDocsWarning,.zDocsCaution,.zDocsTip,.zDocsImportant,[class*="admonition"],[class*="Admonition"],[class*="Notice"],[class*="notice"]');
                if (isAdmonitionIcon || parent) return true;
                // Filter out pure chrome/UI images
                const uiNoise = ['logo', 'icon-', '-icon', 'arrow', 'chevron', 'social-', 'spacer', 'pixel', 'search', 'menu', 'caret', 'close', 'hamburger'];
                if (uiNoise.some(n => filename.includes(n))) return false;
                if (src.includes('/etc.clientlibs/') || src.includes('/clientlibs/')) return false;
                return true;
            }}).map(img => {{
                const src = img.getAttribute('src') || '';
                const alt = img.getAttribute('alt') || '';
                const filename = src.split('/').pop().split('?')[0];
                
                // Check for image breaking issues
                const imgWidth = img.offsetWidth || img.naturalWidth || 0;
                const imgHeight = img.offsetHeight || img.naturalHeight || 0;
                const isOversized = imgWidth > containerWidth;
                const isMissing = !img.offsetWidth || !img.offsetHeight || (img.offsetWidth === 0 && img.offsetHeight === 0);
                const hasNoAlt = !alt || alt.trim().length === 0;
                const brokenIssues = [];
                
                if (isOversized) brokenIssues.push('Image exceeds container width');
                if (isMissing) brokenIssues.push('Image not loading/displaying');
                if (imgHeight > containerWidth * 1.5) brokenIssues.push('Image aspect ratio too tall');
                if (hasNoAlt) brokenIssues.push('Missing alt text');
                
                return {{ src: filename, alt, width: imgWidth, height: imgHeight, 
                         isOversized, isMissing, hasNoAlt,
                         breakingIssues: brokenIssues.length > 0 ? brokenIssues : null }};
            }});

            // Admonition blocks (Note / Warning / Caution / Tip / Important)
            // AEM DITA uses: div.note, div.warning, div.caution, div.tip, div.important
            // Prod (zDocs) uses: [class*="zDocsNote"], [class*="zDocsWarning"], etc.
            const admonitionSelectors = {{
                note:      '.note, [class*="zDocsNote"], [class*="Note--note"], [data-type="note"]',
                warning:   '.warning, [class*="zDocsWarning"], [class*="Note--warning"], [data-type="warning"]',
                caution:   '.caution, [class*="zDocsCaution"], [class*="Note--caution"], [data-type="caution"]',
                tip:       '.tip, [class*="zDocsTip"], [class*="Note--tip"], [data-type="tip"]',
                important: '.important, [class*="zDocsImportant"], [class*="Note--important"], [data-type="important"]',
                danger:    '.danger, [class*="zDocsDanger"], [class*="Note--danger"], [data-type="danger"]',
            }};
            const admonitions = {{}};
            const admonitionTexts = [];
            for (const [type, sel] of Object.entries(admonitionSelectors)) {{
                try {{
                    const els = [...clone.querySelectorAll(sel)];
                    admonitions[type] = els.length;
                    els.forEach(el => {{
                        const txt = (el.innerText || '').trim().replace(/\\s+/g,' ').substring(0, 150);
                        if (txt.length > 5) admonitionTexts.push({{ type, text: txt }});
                    }});
                }} catch(e) {{ admonitions[type] = 0; }}
            }}
            const admonitionTotal = Object.values(admonitions).reduce((a,b) => a+b, 0);
            
            // Icons (SVG inline or icon fonts) — exclude admonition icons already counted
            const iconCount = clone.querySelectorAll('svg:not([class*="admonition"] svg), [class*="icon"]:not([class*="Note"]):not([class*="Admonition"]), .fa').length;

            return {{ 
                title, h1, h2, h3, paras, lists, fullText, bolds, italics, emTexts,
                tableData, imageData, admonitions, admonitionTexts, admonitionTotal,
                sentenceCount: sentences.length,
                bodyTextSnippet, ulCount, iconCount,
                tableBreakingIssues, imageBreakingIssues, textBreakingIssues,
                tags: {{
                    p: paras.length,
                    h2: h2.length,
                    h3: h3.length,
                    li: lists.length,
                    ul: ulCount,
                    em: emTexts.length,
                    img: imageData.length,
                    table: tableData.length,
                    note: admonitions.note || 0,
                    warning: admonitions.warning || 0,
                    caution: admonitions.caution || 0,
                    tip: admonitions.tip || 0,
                    important: admonitions.important || 0,
                }}
            }};
        }}""", is_stage)

        # 4. Debug and Return
        p_count = content.get('tags', {}).get('p', 0)
        full_text_len = len(content.get('fullText', ''))
        if not is_stage and (p_count == 0 or full_text_len < 100):
            snippet = content.get('bodyTextSnippet', '')
            print(f"  ⚠️ No content for {url[:100]} | p={p_count} len={full_text_len} | Body: {snippet[:200]}...")
            
        return content
    except Exception as e:
        print(f"  ❌ Failed to extract: {e}")
        return None


def match_topics(stage_toc, prod_toc):
    def slug(url):
        # /page/topic-name.html -> topic-name
        m = re.search(r'/page/([^/]+?)(?:\.html)?$', url)
        if m: return m.group(1).lower().replace('_','').replace('-','')
        # /bundle/bundle-name/topic-name -> topic-name
        m2 = re.search(r'/bundle/[^/]+/([^/]+?)(?:\.html)?$', url)
        if m2: return m2.group(1).lower().replace('_','').replace('-','')
        # Last part of URL
        path = url.split('/')[-1].split('.')[0]
        return path.lower().replace('_','').replace('-','')

    pairs, used_prod = [], set()
    for s in stage_toc:
        best, best_score = None, 0
        for i, p in enumerate(prod_toc):
            if i in used_prod: continue
            score = max(sim(s['title'], p['title']), sim(slug(s['url']), slug(p['url'])) * 0.9)
            if score > best_score: best_score, best = score, (i, p)
        if best and best_score >= 50:
            used_prod.add(best[0])
            pairs.append({'stage': s, 'prod': best[1], 'matched': True})
        else:
            pairs.append({'stage': s, 'prod': None, 'matched': False})
    for i, p in enumerate(prod_toc):
        if i not in used_prod:
            pairs.append({'stage': None, 'prod': p, 'matched': False})
    return pairs


async def run_deep_validation():
    STAGE_URL = os.environ.get('STAGE_URL','')
    PROD_URL  = os.environ.get('PROD_URL','')
    THRESHOLD = int(os.environ.get('THRESHOLD','80'))
    REPORT_FILENAME = os.environ.get('REPORT_FILENAME') or \
        os.path.join(UI_REPORTS_DIR, f'deep-content-{int(datetime.now().timestamp())}.xlsx')
    TXT_REPORT = REPORT_FILENAME.replace('.xlsx', '.txt')

    if not STAGE_URL or not PROD_URL:
        try:
            with open(TEST_URLS_PATH) as f:
                cfg = json.load(f)
                STAGE_URL = cfg.get('stage','')
                PROD_URL  = cfg.get('production','')
        except: pass

    if not STAGE_URL or not PROD_URL:
        print("❌ Both URLs required"); sys.exit(1)

    print(f"🚀 Full Content Extraction & Validation")
    print(f"   Stage: {STAGE_URL[:60]}")
    print(f"   Prod : {PROD_URL[:60]}\n")

    async with async_playwright() as pw:
        async def setup_ctx(ctx):
            # Do NOT abort CSS/Images if we want visual screenshots
            # await ctx.route("**/*.{png,jpg,jpeg,gif,svg,css,woff,woff2,otf,ttf}", lambda r: r.abort())
            return ctx

        # 1. Main browser for Stage (uses auth)
        s_browser = await pw.chromium.launch(headless=True)
        auth = AUTH_STATE_PATH if os.path.exists(AUTH_STATE_PATH) else None
        s_ctx = await setup_ctx(await s_browser.new_context(storage_state=auth, viewport={'width': 1280, 'height': 800}))
        
        # 2. SEPARATE browser for Production (Incognito/Clean)
        p_browser = await pw.chromium.launch(headless=True)
        p_ctx = await setup_ctx(await p_browser.new_context(viewport={'width': 1280, 'height': 800}))

        # Extract TOC
        print("📋 Extracting TOC from both environments (Separate Browsers)...")
        
        # PROD NORMALIZATION: If URL is a topic page, get bundle root for TOC extraction
        p_toc_url = PROD_URL
        if '/page/' in PROD_URL:
            # e.g. https://documentation.avaya.com/bundle/BundleName/page/Topic.html → strip /page/...
            p_toc_url = PROD_URL.split('/page/')[0]
        elif '.html' in PROD_URL and '/bundle/' in PROD_URL:
            p_toc_url = PROD_URL.rsplit('/', 1)[0]
        
        # STAGE NORMALIZATION: If URL has a specific topic path, get the bundle root
        s_toc_url = STAGE_URL
        if '/page/' in STAGE_URL:
            s_toc_url = STAGE_URL.split('/page/')[0]
        
        print(f"   TOC extraction URLs:")
        print(f"   Stage TOC: {s_toc_url}")
        print(f"   Prod  TOC: {p_toc_url}")
        
        sp, pp = await s_ctx.new_page(), await p_ctx.new_page()
        stage_toc, prod_toc = await asyncio.gather(
            extract_toc(sp, s_toc_url, True),
            extract_toc(pp, p_toc_url, False),
        )
        await sp.close(); await pp.close()
        print(f"   Stage: {len(stage_toc)} | Prod: {len(prod_toc)} topics\n")

        # Match topics
        print("🔗 Matching topics...")
        pairs = match_topics(stage_toc, prod_toc)
        matched = [p for p in pairs if p['matched']]
        only_stage = [p for p in pairs if not p['matched'] and p['stage']]
        only_prod = [p for p in pairs if not p['matched'] and p['prod']]
        print(f"   Matched: {len(matched)} | Only Stage: {len(only_stage)} | Only Prod: {len(only_prod)}\n")

        # Extract content for each matched pair
        print(f"📄 Extracting full content from {len(matched)} matched topics...\n")
        sem = asyncio.Semaphore(CONCURRENCY)
        
        async def compare(pair, idx):
            async with sem:
                s_item, p_item = pair['stage'], pair['prod']
                print(f"  [{idx}/{len(matched)}] {s_item['title'][:60]}")
                
                # Shared context for Stage (to keep auth state)
                sg = await s_ctx.new_page()
                
                # FRESH browser & context for Production (as requested: new browser incognito)
                pg_browser = await pw.chromium.launch(headless=True)
                pg_ctx = await setup_ctx(await pg_browser.new_context(viewport={'width': 1280, 'height': 800}))
                pg = await pg_ctx.new_page()
                
                try:
                    s_content, p_content = await asyncio.gather(
                        get_full_content(sg, s_item['url'], is_stage=True),
                        get_full_content(pg, p_item['url'], is_stage=False),
                    )
                except Exception as e:
                    print(f"  ⚠️ Content extraction failed: {e}")
                    s_content, p_content = None, None

                await sg.close()
                await pg.close()
                await pg_ctx.close() 
                await pg_browser.close() # Close fresh browser

                if not s_content or not p_content:
                    return None

                # Compare content
                title_sim = sim(s_content.get('title',''), p_content.get('title',''))
                text_sim  = sim(s_content.get('fullText','')[:8000], p_content.get('fullText','')[:8000])
                
                # Granular Heading Comparison - fuzzy match
                s_h_list = s_content.get('h2', []) + s_content.get('h3', [])
                p_h_list = p_content.get('h2', []) + p_content.get('h3', [])
                
                def fuzzy_matched(text, candidates, threshold=85):
                    """Return True if text closely matches any candidate."""
                    t = text.lower().strip()
                    for c in candidates:
                        if sim(t, c.lower().strip()) >= threshold:
                            return True
                    return False
                
                def bidirectional_match(list_a, list_b, threshold=80, truncate=None):
                    """
                    Bidirectional fuzzy matching: finds items that exist in both lists.
                    Returns: (matched_items_from_a, unmatched_from_a, unmatched_from_b)
                    """
                    matched_a = []
                    unmatched_a = []
                    unmatched_b = list(list_b)  # Start with all of B
                    
                    for item_a in list_a:
                        test_a = item_a[:truncate].lower().strip() if truncate else item_a.lower().strip()
                        best_match = None
                        best_score = 0
                        
                        for idx_b, item_b in enumerate(unmatched_b):
                            test_b = item_b[:truncate].lower().strip() if truncate else item_b.lower().strip()
                            score = sim(test_a, test_b)
                            if score > best_score:
                                best_score = score
                                best_match = idx_b
                        
                        if best_score >= threshold and best_match is not None:
                            matched_a.append(item_a)
                            unmatched_b.pop(best_match)  # Remove matched item from B
                        else:
                            unmatched_a.append(item_a)
                    
                    return matched_a, unmatched_a, unmatched_b
                
                missing_h = [h for h in p_h_list if not fuzzy_matched(h, s_h_list)]
                extra_h   = [h for h in s_h_list if not fuzzy_matched(h, p_h_list)]
                
                h_sim = sim(' '.join(s_h_list), ' '.join(p_h_list))
                overall = int((title_sim*0.1 + h_sim*0.3 + text_sim*0.6))

                # Detailed Discrepancies
                discrepancies = []
                
                if missing_h: discrepancies.append(f"MISSING HEADINGS: {len(missing_h)} items. First: \"{missing_h[0]}\"")
                if extra_h:   discrepancies.append(f"EXTRA HEADINGS IN STAGE: {len(extra_h)} items. First: \"{extra_h[0]}\"")
                
                # Sentence Count Validation
                s_sent = s_content.get('sentenceCount', 0)
                p_sent = p_content.get('sentenceCount', 0)
                if abs(s_sent - p_sent) > 5:
                    discrepancies.append(f"SENTENCE COUNT: Mismatch (S:{s_sent} vs P:{p_sent})")
                
                # Formatting (Bold/Italic) — fuzzy match
                missing_bold = [b for b in p_content.get('bolds', []) if not fuzzy_matched(b, s_content.get('bolds', []))]
                if missing_bold: discrepancies.append(f"MISSING BOLD TEXT: \"{missing_bold[0][:50]}\"")
                
                missing_ital = [i for i in p_content.get('italics', []) if not fuzzy_matched(i, s_content.get('italics', []))]
                if missing_ital: discrepancies.append(f"MISSING ITALIC TEXT: \"{missing_ital[0][:50]}\"")

                # Versions
                s_ver = s_content.get('version', 'N/A')
                p_ver = p_content.get('version', 'N/A')
                if s_ver != p_ver:
                    if s_ver == 'N/A' or not s_ver: discrepancies.append("VERSION: Not found in Stage")
                    elif p_ver == 'N/A' or not p_ver: discrepancies.append("VERSION: Not found in Prod")
                    else: discrepancies.append(f"VERSION: Mismatch (S:{s_ver} vs P:{p_ver})")

                # Tags mismatch — only report if count diff is significant (>2 for icons)
                s_tags, p_tags = s_content.get('tags', {}), p_content.get('tags', {})
                tag_map = {'p': ('Paragraphs', 1), 'h2': ('H2 Headings', 1), 'h3': ('H3 Headings', 1),
                           'li': ('List Items', 2), 'ul': ('Lists (ul/ol)', 1),
                           'em': ('Emphasized Text', 2), 'img': ('Images', 1), 'table': ('Tables', 1)}
                # Skip icon count — icons often differ between stage/prod rendering
                for tag, (label, tolerance) in tag_map.items():
                    sc, pc = s_tags.get(tag, 0), p_tags.get(tag, 0)
                    if sc < pc - tolerance: discrepancies.append(f"MISSING {label.upper()}: Stage={sc}, Prod={pc} (diff: {pc-sc})")
                    elif sc > pc + tolerance: discrepancies.append(f"EXTRA {label.upper()} IN STAGE: Stage={sc}, Prod={pc} (diff: {sc-pc})")

                # Admonition (Note/Warning/Caution/Tip/Important) comparison
                s_admon = s_content.get('admonitions', {})
                p_admon = p_content.get('admonitions', {})
                for atype in ['note', 'warning', 'caution', 'tip', 'important', 'danger']:
                    sc, pc = s_admon.get(atype, 0), p_admon.get(atype, 0)
                    if sc < pc:
                        discrepancies.append(f"MISSING {atype.upper()} ADMONITION: Stage={sc}, Prod={pc} (diff: {pc-sc})")
                    elif sc > pc:
                        discrepancies.append(f"EXTRA {atype.upper()} IN STAGE: Stage={sc}, Prod={pc} (diff: {sc-pc})")
                
                # Admonition text content fuzzy match
                s_admon_texts = [a['text'] for a in s_content.get('admonitionTexts', [])]
                p_admon_texts = [a['text'] for a in p_content.get('admonitionTexts', [])]
                missing_admon_text = [a for a in p_content.get('admonitionTexts', [])
                                      if not fuzzy_matched(a['text'][:150], s_admon_texts, threshold=75)]
                if missing_admon_text:
                    first = missing_admon_text[0]
                    discrepancies.append(f"MISSING {first['type'].upper()} CONTENT: \"{first['text'][:80]}\"" +
                                         (f" (+{len(missing_admon_text)-1} more)" if len(missing_admon_text) > 1 else ""))

                # Table content comparison (row-level) + breaking issues
                s_tables = s_content.get('tableData', [])
                p_tables = p_content.get('tableData', [])
                if len(s_tables) != len(p_tables):
                    discrepancies.append(f"TABLE COUNT: Stage={len(s_tables)} vs Prod={len(p_tables)}")
                else:
                    for ti, (st, pt) in enumerate(zip(s_tables, p_tables)):
                        if isinstance(st, dict) and isinstance(pt, dict):
                            if abs(st.get('rows',0) - pt.get('rows',0)) > 1:
                                discrepancies.append(f"TABLE {ti+1} ROW COUNT: Stage={st.get('rows')} vs Prod={pt.get('rows')}")
                            if sim(st.get('content',''), pt.get('content','')) < 70:
                                discrepancies.append(f"TABLE {ti+1} CONTENT MISMATCH")
                
                # Table breaking issues comparison
                s_table_breaks = sum(1 for t in s_tables if isinstance(t, dict) and t.get('breakingIssues'))
                p_table_breaks = sum(1 for t in p_tables if isinstance(t, dict) and t.get('breakingIssues'))
                if s_table_breaks > p_table_breaks + 1:
                    discrepancies.append(f"TABLE BREAKING: Stage has {s_table_breaks} tables with layout issues vs Prod {p_table_breaks}")
                elif p_table_breaks > s_table_breaks:
                    discrepancies.append(f"TABLE BREAKING IN PROD: {p_table_breaks} tables with layout issues")
                


                # Image comparison (by filename) + breaking issues
                s_imgs = set(i.get('src','') if isinstance(i,dict) else i for i in s_content.get('imageData', []))
                p_imgs = set(i.get('src','') if isinstance(i,dict) else i for i in p_content.get('imageData', []))
                missing_imgs = list(p_imgs - s_imgs)
                extra_imgs   = list(s_imgs - p_imgs)
                if missing_imgs:
                    discrepancies.append(f"MISSING IMAGES IN STAGE: {', '.join(missing_imgs[:3])}")
                if extra_imgs:
                    discrepancies.append(f"EXTRA IMAGES IN STAGE: {', '.join(extra_imgs[:3])}")
                
                # Image breaking issues comparison
                s_img_breaks = sum(1 for img in s_content.get('imageData', []) if isinstance(img, dict) and img.get('breakingIssues'))
                p_img_breaks = sum(1 for img in p_content.get('imageData', []) if isinstance(img, dict) and img.get('breakingIssues'))
                if s_img_breaks > p_img_breaks + 1:
                    discrepancies.append(f"IMAGE BREAKING: Stage has {s_img_breaks} images with layout issues vs Prod {p_img_breaks}")
                elif p_img_breaks > s_img_breaks:
                    discrepancies.append(f"IMAGE BREAKING IN PROD: {p_img_breaks} images with sizing/display issues")
                
                # Text breaking issues comparison
                s_text_breaks = s_content.get('textBreakingIssues', []) or []
                p_text_breaks = p_content.get('textBreakingIssues', []) or []
                if len(s_text_breaks) > len(p_text_breaks):
                    for issue in s_text_breaks[:2]:  # Report first 2 issues from stage
                        if issue not in p_text_breaks:
                            discrepancies.append(f"TEXT BREAKING (Stage): {issue[:80]}")
                elif len(p_text_breaks) > len(s_text_breaks):
                    for issue in p_text_breaks[:2]:  # Report first 2 issues from prod
                        if issue not in s_text_breaks:
                            discrepancies.append(f"TEXT BREAKING (Prod): {issue[:80]}")

                # Emphasized text (em) — bidirectional fuzzy match
                _, missing_em, extra_em = bidirectional_match(s_content.get('emTexts',[]), p_content.get('emTexts',[]), threshold=80)
                if missing_em:
                    discrepancies.append(f"MISSING EMPHASIZED TEXT: \"{missing_em[0][:60]}\"" + (f" (+{len(missing_em)-1} more)" if len(missing_em)>1 else ""))

                # Paragraph / Body text — bidirectional fuzzy match (threshold 80 to handle minor diffs)
                s_paras = s_content.get('paras', [])
                p_paras = p_content.get('paras', [])
                
                _, missing_p, extra_p = bidirectional_match(s_paras, p_paras, threshold=80, truncate=200)
                
                if missing_p:
                    discrepancies.append(f"MISSING BODY TEXT ({len(missing_p)} items): \"{missing_p[0][:80]}\"")
                if extra_p:
                    discrepancies.append(f"EXTRA BODY TEXT IN PROD ({len(extra_p)} items): \"{extra_p[0][:80]}\"")
                
                # List item Matching — bidirectional fuzzy match
                s_li_list = s_content.get('lists', [])
                p_li_list = p_content.get('lists', [])
                _, missing_li, extra_li = bidirectional_match(s_li_list, p_li_list, threshold=80, truncate=150)
                if missing_li: discrepancies.append(f"MISSING LIST CONTENT ({len(missing_li)} items): \"{missing_li[0][:80]}\"")
                if extra_li: discrepancies.append(f"EXTRA LIST CONTENT ({len(extra_li)} items): \"{extra_li[0][:80]}\"")

                status = 'MATCH' if overall >= THRESHOLD else ('PARTIAL' if overall >= 60 else 'MISMATCH')

                # Group findings for better columns
                format_f   = ' | '.join([d for d in discrepancies if "BOLD" in d or "ITALIC" in d])
                asset_f    = ' | '.join([d for d in discrepancies if "IMAGE" in d or "TABLE" in d])
                version_f  = ' | '.join([d for d in discrepancies if "VERSION" in d])

                return {
                    'Topic': s_item['title'], 
                    'Status': status,
                    'Overall %': overall,
                    'Mismatch Summary': ' | '.join(discrepancies) if discrepancies else 'MATCH',
                    'Formatting Issues': format_f or 'MATCH',
                    'Asset Issues': asset_f or 'MATCH',
                    'Version Status': version_f or 'MATCH',
                    'S-URL': s_item['url'], 
                    'P-URL': p_item['url'],
                    'S-Stats': f"P:{s_tags.get('p')}|I:{s_tags.get('img')}|T:{s_tags.get('table')}|Note:{s_tags.get('note',0)}|Warn:{s_tags.get('warning',0)}",
                    'P-Stats': f"P:{p_tags.get('p')}|I:{p_tags.get('img')}|T:{p_tags.get('table')}|Note:{p_tags.get('note',0)}|Warn:{p_tags.get('warning',0)}"
                }

        results = await asyncio.gather(*[compare(p, i+1) for i, p in enumerate(matched)])
        rows = [r for r in results if r]
        
        # Add unmatched stage topics
        for p in only_stage:
            rows.append({
                'Topic': p['stage']['title'], 'Status': 'MISSING IN PROD', 'Overall %': 0,
                'Mismatch Summary': 'Topic not found in Production TOC',
                'Formatting Issues': 'N/A', 'Asset Issues': 'N/A', 'Version Status': 'N/A',
                'S-URL': p['stage']['url'], 'P-URL': 'N/A', 'S-Stats': 'N/A', 'P-Stats': 'N/A'
            })
        # Add unmatched prod topics
        for p in only_prod:
            rows.append({
                'Topic': p['prod']['title'], 'Status': 'MISSING IN STAGE', 'Overall %': 0,
                'Mismatch Summary': 'Topic not found in Stage TOC',
                'Formatting Issues': 'N/A', 'Asset Issues': 'N/A', 'Version Status': 'N/A',
                'S-URL': 'N/A', 'P-URL': p['prod']['url'], 'S-Stats': 'N/A', 'P-Stats': 'N/A'
            })
            
        await s_browser.close()
        await p_browser.close()

    # Process missing topics for the report
    missing_stage = [{'topic': p['prod']['title'], 'type': 'Topic', 'prod': p['prod']['url'], 'stage': 'NOT FOUND'} for p in only_prod]
    missing_prod = [{'topic': p['stage']['title'], 'type': 'Topic', 'stage': p['stage']['url'], 'prod': 'NOT FOUND'} for p in only_stage]

    # Generate TXT Report
    print(f"\n📝 Generating reports...")
    txt_lines = [
        "="*100,
        "FULL CONTENT EXTRACTION & VALIDATION REPORT",
        "="*100,
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Stage: {STAGE_URL}",
        f"Prod : {PROD_URL}",
        f"Threshold: {THRESHOLD}%",
        "",
        "─"*100,
        "SUMMARY",
        "─"*100,
        f"Total Stage Topics   : {len(stage_toc)}",
        f"Total Prod Topics    : {len(prod_toc)}",
        f"Matched Pairs        : {len(matched)}",
        f"Only in Stage        : {len(only_stage)}",
        f"Only in Prod         : {len(only_prod)}",
        "",
    ]

    if rows:
        match_c = sum(1 for r in rows if r['Status'] == 'MATCH')
        partial_c = sum(1 for r in rows if r['Status'] == 'PARTIAL')
        mismatch_c = sum(1 for r in rows if r['Status'] == 'MISMATCH')
        missing_c = sum(1 for r in rows if 'MISSING' in r.get('Status',''))
        overall = int((match_c + partial_c*0.5) / len(rows) * 100) if rows else 0
    else:
        match_c = partial_c = mismatch_c = missing_c = overall = 0

    txt_lines.extend([
        f"Content Match        : {match_c}",
        f"Content Partial      : {partial_c}",
        f"Content Mismatch     : {mismatch_c}",
        f"Missing Topics       : {missing_c}",
        f"Overall Score        : {overall}%",
        "",
        "Topic-wise Similarity Summary:",
    ])
    for r in rows:
        txt_lines.append(f"  • {r['Topic'][:50].ljust(50)} : {r['Overall %']}% ({r['Status']})")

    if rows:
        txt_lines.extend([
            "",
            "─"*100,
            "DETAILED COMPARISON",
            "─"*100,
            "",
        ])

        for row in rows:
            txt_lines.extend([
                f"Topic: {row['Topic']}",
                f"  Status: {row['Status']} (Score: {row['Overall %']}%)",
                f"  Mismatch Reason: {row['Mismatch Summary']}",
                f"  Formatting Match: {row['Formatting Issues']}",
                f"  Asset Findings: {row['Asset Issues']}",
                f"  Version Status: {row['Version Status']}",
                f"  Stage URL: {row['S-URL']}",
                f"  Prod URL : {row['P-URL']}",
                f"  Content Stats (Stage vs Prod):",
                f"    Stats: {row['S-Stats']} vs {row['P-Stats']}",
                "",
            ])

    # Missing content
    if missing_stage or missing_prod:
        txt_lines.extend([
            "─"*100,
            "MISSING TOPICS ANALYSIS",
            "─"*100,
            "",
        ])

    if missing_stage:
        txt_lines.extend([
            "📌 MISSING IN STAGE (Available in Prod):",
            "",
        ])
        for m in missing_stage:
            txt_lines.append(f"  • {m['topic']}")
            txt_lines.append(f"    Prod URL: {m['prod']}")

    if missing_prod:
        txt_lines.extend([
            "",
            "📌 MISSING IN PROD (Available in Stage):",
            "",
        ])
        for m in missing_prod:
            txt_lines.append(f"  • {m['topic']}")
            txt_lines.append(f"    Stage URL: {m['stage']}")

    txt_lines.extend([
        "",
        "─"*100,
        "END OF REPORT",
        "─"*100,
    ])

    with open(TXT_REPORT, 'w') as f:
        f.write('\n'.join(txt_lines))

    # Generate XLSX
    summary_data = [
        ['Full Content Validation Report'],
        ['Date', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['Stage URL', STAGE_URL],
        ['Prod URL', PROD_URL],
        ['Threshold', f"{THRESHOLD}%"],
        [],
        ['Metric', 'Value'],
        ['Stage Topics', len(stage_toc)],
        ['Prod Topics', len(prod_toc)],
        ['Matched Pairs', len(matched)],
        ['Only in Stage', len(only_stage)],
        ['Only in Prod', len(only_prod)],
        [],
        ['Content Match', match_c],
        ['Content Partial', partial_c],
        ['Content Mismatch', mismatch_c],
        ['Missing Topics', missing_c],
        ['Overall Score', f"{overall}%"],
        [],
        ['Topic-wise Similarity Summary'],
        ['Topic Name', 'Match %', 'Status']
    ]
    
    for r in rows:
        summary_data.append([r['Topic'], f"{r['Overall %']}%", r['Status']])

    with pd.ExcelWriter(REPORT_FILENAME, engine='openpyxl') as w:
        pd.DataFrame(summary_data).to_excel(w, sheet_name='Summary', header=False, index=False)
        if rows:
            df = pd.DataFrame(rows)
            df.to_excel(w, sheet_name='Content Comparison', index=False)
            
            # Formatting
            ws = w.sheets['Content Comparison']
            
            # Widths
            ws.column_dimensions['A'].width = 40 # Topic
            ws.column_dimensions['B'].width = 12 # Status
            ws.column_dimensions['C'].width = 10 # Overall %
            ws.column_dimensions['D'].width = 50 # Mismatch Summary
            ws.column_dimensions['E'].width = 35 # Formatting Issues
            ws.column_dimensions['F'].width = 35 # Asset Issues
            ws.column_dimensions['G'].width = 25 # Version Status
            ws.column_dimensions['H'].width = 40 # S-URL
            ws.column_dimensions['I'].width = 40 # P-URL
            
            # Color coding
            from openpyxl.styles import PatternFill
            green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            orange = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            
            for row_idx in range(2, len(rows) + 2):
                status = ws.cell(row=row_idx, column=2).value
                if status == 'MATCH': ws.cell(row=row_idx, column=2).fill = green
                elif status in ['MISMATCH', 'MISSING IN PROD', 'MISSING IN STAGE']: ws.cell(row=row_idx, column=2).fill = red
                elif status == 'PARTIAL': ws.cell(row=row_idx, column=2).fill = orange
                
                # Highlight mismatch cells (Col 4 to 7 - Discrepancy columns)
                for col_idx in range(4, 8):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val and val != 'MATCH' and val != 'N/A':
                        ws.cell(row=row_idx, column=col_idx).fill = red

    print(f"✅ Reports saved:")
    print(f"   Excel: {REPORT_FILENAME}")
    print(f"   TXT  : {TXT_REPORT}")

    print(f"::RESULTS::{json.dumps({'overall':overall if rows else 0,'match':match_c if rows else 0,'partial':partial_c if rows else 0,'mismatch':mismatch_c if rows else 0,'stage_total':len(stage_toc),'prod_total':len(prod_toc),'matched':len(matched),'only_stage':len(only_stage),'only_prod':len(only_prod),'missing_stage_count':len(missing_stage),'missing_prod_count':len(missing_prod),'reportFile':os.path.basename(REPORT_FILENAME),'txtFile':os.path.basename(TXT_REPORT)})}")


if __name__ == "__main__":
    asyncio.run(run_deep_validation())

