import os
import sys
import json
import asyncio
from datetime import datetime
from urllib.parse import urlparse
import pandas as pd
from playwright.async_api import async_playwright

# ── Config ────────────────────────────────────────────────────────────
UI_REPORTS_DIR = os.path.join(os.getcwd(), '.ui_reports')
os.makedirs(UI_REPORTS_DIR, exist_ok=True)

AUTH_STATE_PATH = os.path.join(os.getcwd(), 'auth-sessions', 'storage-state.json')

START_URL = os.environ.get('START_URL', 'https://publish-p181473-e1910385.adobeaemcloud.com/en-us/home/bundle/')
REPORT_FILENAME = os.environ.get('REPORT_FILENAME') or os.path.join(
    UI_REPORTS_DIR, f'extra-products-{int(datetime.now().timestamp())}.xlsx'
)

parsed = urlparse(START_URL)
BASE = f"{parsed.scheme}://{parsed.netloc}"

print(f"🚀 Starting Extra Products Discovery")
print(f"   URL: {START_URL}")

async def handle_cookies(page):
    try:
        for sel in ['#onetrust-accept-btn-handler', '#btn-accept-all', 'button:has-text("Accept")', '.cookie-accept']:
            if await page.locator(sel).is_visible(timeout=1500):
                await page.click(sel)
                await page.wait_for_timeout(500)
                break
    except:
        pass

async def run():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(
            storage_state=AUTH_STATE_PATH if os.path.exists(AUTH_STATE_PATH) else None
        )
        
        # 10x Speedup Optimization: Block unnecessary requests (images, CSS, webfonts, trackers)
        async def block_unnecessary(route):
            req = route.request
            if req.is_navigation_request():
                await route.continue_()
                return
            
            res_type = req.resource_type
            url = req.url.lower()
            if res_type in ['image', 'media', 'font', 'stylesheet'] or \
               any(k in url for k in ['analytics', 'optimizely', 'dynatrace', 'tracking', 'google-analytics', 'metrics']) or \
               (res_type == 'script' and 'adobe' in url and 'adobeaemcloud' not in url):
                await route.abort()
            else:
                await route.continue_()
                
        await ctx.route("**/*", block_unnecessary)
        
        page = await ctx.new_page()

        print(f"\n🔍 Loading page...")
        try:
            # 10x Speedup Optimization: Wait only until DOMContentLoaded, avoiding networkidle wait times
            await page.goto(START_URL, wait_until='domcontentloaded', timeout=30000)
        except Exception as e:
            print(f"   ⚠️ Load warning: {e}")

        await handle_cookies(page)
        await page.wait_for_timeout(500)

        title = await page.title()
        print(f"   ✅ Page loaded: {title}")

        # Extract main page / product heading from the bundle page
        heading = ""
        try:
            heading = await page.locator('h1, .bundle-hero__title, .hero__title, .topic-renderer__title').first.inner_text(timeout=2000)
            heading = heading.strip()
        except:
            heading = title
        print(f"   👤 Product Heading: {heading}")

        # Extract all sitemap and product links from the page
        links_data = await page.evaluate(f'''() => {{
            const base = "{BASE}";
            const results = [];
            const seen = new Set();
            
            // 10x Speedup Optimization: Query sitemap links natively in C++ once and store in a Set for O(1) membership test
            const sitemapSet = new Set();
            try {{
                document.querySelectorAll([
                    '.sitemap a[href]',
                    '.cmp-sitemap a[href]',
                    '.sitemap-list a[href]',
                    '.sitemap-tree a[href]',
                    '.sitemap-container a[href]',
                    '.sitemap-tree-node a[href]',
                    'ul.sitemap a[href]'
                ].join(',')).forEach(a => sitemapSet.add(a));
            }} catch(e) {{}}

            document.querySelectorAll('a[href]').forEach(a => {{
                const href = a.getAttribute('href') || '';
                if (!href || href.startsWith('#') || href.startsWith('javascript') || href.startsWith('mailto:') || href.startsWith('tel:')) return;
                
                let fullUrl = href.startsWith('http') ? href : base + (href.startsWith('/') ? href : '/' + href);
                fullUrl = fullUrl.split('?')[0].split('#')[0];
                
                if (!fullUrl.startsWith(base)) return;
                if (seen.has(fullUrl)) return;
                seen.add(fullUrl);
                
                let text = a.innerText.trim() || a.getAttribute('title') || '';
                if (!text) {{
                    const h = a.querySelector('h1,h2,h3,h4,h5');
                    if (h) text = h.innerText.trim();
                }}
                if (!text) {{
                    text = fullUrl.split('/').pop().replace('.html','').replace(/-/g,' ');
                    text = text.charAt(0).toUpperCase() + text.slice(1);
                }}
                
                // Exclude boilerplate utility links to keep the report super clean
                const lowerText = text.toLowerCase();
                if (lowerText.includes('privacy') || lowerText.includes('legal') || lowerText.includes('cookie') || 
                    lowerText.includes('terms of') || lowerText.includes('contact us') || lowerText.includes('support') ||
                    lowerText.includes('sign in') || lowerText.includes('login') || lowerText === 'home' || 
                    lowerText === 'avaya' || lowerText === 'search') {{
                    return;
                }}

                // Classify as Sitemap Link or Product Link using the fast O(1) Set lookup
                let linkType = 'Product Link';
                const lowerUrl = fullUrl.toLowerCase();
                if (sitemapSet.has(a) || lowerUrl.includes('sitemap') || lowerText.includes('sitemap') || lowerText.includes('sitemap link')) {{
                    linkType = 'Sitemap Link';
                }}

                results.push({{ title: text, url: fullUrl, type: linkType }});
            }});
            return results;
        }}''')

        await browser.close()

        print(f"   📦 Found {len(links_data)} discovered links")
        if links_data:
            for item in links_data[:3]:
                print(f"   - [{item['type']}] {item['title'][:60]}")

        # ── Excel Report ─────────────────────────────────────────────
        total = len(links_data)
        sitemap_cnt = len([x for x in links_data if x.get('type') == 'Sitemap Link'])
        product_cnt = len([x for x in links_data if x.get('type') == 'Product Link'])
        
        summary_data = [
            ['Extra Products Discovery Report'],
            [''],
            ['Date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Product Heading', heading],
            ['Source URL', START_URL],
            [''],
            ['── Results ──'],
            ['Total Links Discovered', total],
            ['Sitemap Links Found', sitemap_cnt],
            ['Product Links Found', product_cnt],
        ]

        os.makedirs(os.path.dirname(REPORT_FILENAME), exist_ok=True)

        with pd.ExcelWriter(REPORT_FILENAME, engine='openpyxl') as writer:
            from openpyxl.styles import Font, PatternFill, Alignment

            pd.DataFrame(summary_data).to_excel(writer, sheet_name='Summary', header=False, index=False)

            if links_data:
                products_df = pd.DataFrame([{
                    '#': i + 1,
                    'Product Heading': heading,
                    'Link Title': item.get('title', ''),
                    'Link URL':   item.get('url', ''),
                    'Classification': item.get('type', 'Product Link'),
                } for i, item in enumerate(links_data)])
                products_df.to_excel(writer, sheet_name='Products', index=False)

                ws = writer.sheets['Products']
                red_fill = PatternFill(start_color='DA291C', end_color='DA291C', fill_type='solid')
                for cell in ws[1]:
                    cell.fill = red_fill
                    cell.font = Font(color='FFFFFF', bold=True)
                    cell.alignment = Alignment(horizontal='center')
                ws.column_dimensions['A'].width = 6
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 50
                ws.column_dimensions['D'].width = 80
                ws.column_dimensions['E'].width = 20
            else:
                pd.DataFrame([{'Message': 'No links found on the page'}]).to_excel(
                    writer, sheet_name='Products', index=False)

        results = {'overall': total, 'bundle_pages': 1, 'visited': 1}
        print(f"\n✅ Report saved: {REPORT_FILENAME}")
        print(f"::RESULTS::{json.dumps(results)}")

if __name__ == "__main__":
    asyncio.run(run())

